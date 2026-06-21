#!/usr/bin/env python3
"""
Complex Project Timeline & Gantt Chart Generator
Multi-dimensional with formula interactions, conditional formatting, data validation, and rich styling.
"""
import openpyxl, os, datetime
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle,
    numbers, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule, ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "03_Business_Management")
OUT_FILE = os.path.join(OUT, "03_Project_Timeline.xlsx")
SITE = "https://planoranest.com/"
TD = datetime.date.today().strftime("%B %d, %Y")
YR = datetime.datetime.now().year

# ═══════════════════════════════════════════
# COLOR SYSTEM
# ═══════════════════════════════════════════
NAVY      = "1B2A4A"
DARK_BLUE = "1F4E79"
STEEL     = "2E75B6"
ACCENT    = "4472C4"
LIGHT_BG  = "F2F7FB"
WHITE_BG  = "FFFFFF"
DARK_BG   = "E8F0FE"
GREEN     = "1B7A3D"
RED       = "C0392B"
ORANGE    = "E67E22"
YELLOW    = "F1C40F"
PURPLE    = "7D3C98"
GRAY      = "95A5A6"
DARK_TEXT  = "1A1A1A"
MED_TEXT   = "555555"

# ═══════════════════════════════════════════
# STYLES
# ═══════════════════════════════════════════
thin_border = Border(
    left=Side('thin', color='D5D8DC'), right=Side('thin', color='D5D8DC'),
    top=Side('thin', color='D5D8DC'), bottom=Side('thin', color='D5D8DC')
)
header_border = Border(
    left=Side('thin', color='1B2A4A'), right=Side('thin', color='1B2A4A'),
    top=Side('thin', color='1B2A4A'), bottom=Side('medium', color='1B2A4A')
)

hdr_fill   = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
hdr_font   = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
title_font = Font(name='Calibri', size=20, bold=True, color=NAVY)
sub_font   = Font(name='Calibri', size=11, italic=True, color=MED_TEXT)
section_font = Font(name='Calibri', size=13, bold=True, color=DARK_BLUE)
data_font  = Font(name='Calibri', size=10, color=DARK_TEXT)
link_font  = Font(name='Calibri', size=10, color='0563C1', underline='single')
small_font = Font(name='Calibri', size=8, color='999999')
note_font  = Font(name='Calibri', size=9, italic=True, color=MED_TEXT)
kpi_val_font = Font(name='Calibri', size=24, bold=True, color=DARK_BLUE)
kpi_lbl_font = Font(name='Calibri', size=10, color=MED_TEXT)

alt_fill   = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
green_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
red_fill   = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
yellow_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
orange_fill = PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid')

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align  = Alignment(horizontal='right', vertical='center')

def apply_cell(ws, row, col, value, font=data_font, fill=None, alignment=left_align, border=thin_border, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font; cell.alignment = alignment; cell.border = border
    if fill: cell.fill = fill
    if number_format: cell.number_format = number_format
    return cell

def apply_header(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = apply_cell(ws, row, start_col+i, h, font=hdr_font, fill=hdr_fill, alignment=center_align, border=header_border)
    ws.row_dimensions[row].height = 32

def apply_data_row(ws, row, values, start_col=1, fonts=None, fills=None, aligns=None, formats=None):
    for i, v in enumerate(values):
        f = fonts[i] if fonts and i < len(fonts) else data_font
        fl = fills[i] if fills and i < len(fills) else (alt_fill if row % 2 == 0 else None)
        a = aligns[i] if aligns and i < len(aligns) else center_align
        nf = formats[i] if formats and i < len(formats) else None
        apply_cell(ws, row, start_col+i, v, font=f, fill=fl, alignment=a, border=thin_border, number_format=nf)

def set_col_widths(ws, widths, start_col=1):
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col+i)].width = w

def merge_title(ws, row, col_start, col_end, text, font=title_font, height=40):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    apply_cell(ws, row, col_start, text, font=font, alignment=left_align, border=Border())
    ws.row_dimensions[row].height = height

def merge_subtitle(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    apply_cell(ws, row, col_start, text, font=sub_font, alignment=left_align, border=Border())

def section_header(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    apply_cell(ws, row, col_start, text, font=section_font, fill=PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid'), alignment=left_align, border=Border())
    ws.row_dimensions[row].height = 28

def merge(ws, r1, c1, r2, c2, text, font=data_font, fill=None):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    apply_cell(ws, r1, c1, text, font=font, fill=fill, alignment=center_align, border=thin_border)

# ═══════════════════════════════════════════
# BUILD WORKBOOK
# ═══════════════════════════════════════════

wb = openpyxl.Workbook()
wb.remove(wb.active)

print("=" * 60)
print("Building Complex Project Timeline & Gantt Chart...")
print("=" * 60)

# ─────────────────────────────────────────
# SHEET 1: PROJECT DASHBOARD
# ─────────────────────────────────────────
ws_dash = wb.create_sheet("📊 Dashboard")
ws_dash.sheet_properties.tabColor = NAVY
set_col_widths(ws_dash, [3, 28, 16, 16, 16, 16, 16, 3])

# -- Title --
merge_title(ws_dash, 1, 2, 7, "📊 Project Timeline Dashboard")
merge_subtitle(ws_dash, 2, 2, 7, "Real-time project metrics, progress tracking & resource overview  |  PlanoraNest.com")

# -- KPI Row --
kpi_row = 4
kpi_cards = [
    ("A", "B", "C", "D", "E", "F"),
    ("OVERALL PROGRESS", "TOTAL TASKS", "COMPLETED", "BUDGET SPENT", "BUDGET REMAINING", "DAYS REMAINING"),
]
# Merge KPI cells as cards
kpi_cols = [2, 3, 4, 5, 6, 7]
kpi_widths = [28, 16, 16, 16, 16, 16]
kpi_labels = ["Overall Progress", "Total Tasks", "Completed", "Budget Spent", "Budget Remaining", "Days Remaining"]

# KPI card styling
kpi_card_fill = PatternFill(start_color=WHITE_BG, end_color=WHITE_BG, fill_type='solid')
kpi_card_border = Border(
    left=Side('medium', color=STEEL), right=Side('thin', color='E5E7E9'),
    top=Side('thin', color='E5E7E9'), bottom=Side('thin', color='E5E7E9')
)

for idx, (col, lbl) in enumerate(zip(kpi_cols, kpi_labels)):
    c = ws_dash.cell(row=kpi_row, column=col, value=lbl)
    c.font = kpi_lbl_font; c.fill = kpi_card_fill; c.alignment = center_align
    c.border = kpi_card_border
    set_col_widths(ws_dash, [kpi_widths[idx]], col)
ws_dash.row_dimensions[kpi_row].height = 22

# KPI values row
kpi_vals_row = kpi_row + 1
kpi_formulas = [
    '=COUNTA(\'📋 Task Tracker\'!B7:B106)',
    '=COUNTIF(\'📋 Task Tracker\'!I7:I106,"Complete")',
    '=\'📋 Task Tracker\'!F7+\'📋 Task Tracker\'!F8+\'📋 Task Tracker\'!F9',
    '=\'📋 Task Tracker\'!G7+\'📋 Task Tracker\'!G8+\'📋 Task Tracker\'!G9',
]
for idx, (col, lbl) in enumerate(zip(kpi_cols, kpi_labels)):
    if idx == 0:
        val = '=COUNTIF(\'📋 Task Tracker\'!J7:J106,"Complete")/MAX(1,COUNTA(\'📋 Task Tracker\'!B7:B106))'
        apply_cell(ws_dash, kpi_vals_row, col, val, font=kpi_val_font, fill=kpi_card_fill,
                   alignment=center_align, border=kpi_card_border, number_format='0%')
    elif idx == 1:
        val = '=COUNTA(\'📋 Task Tracker\'!B7:B106)'
        apply_cell(ws_dash, kpi_vals_row, col, val, font=kpi_val_font, fill=kpi_card_fill,
                   alignment=center_align, border=kpi_card_border)
    elif idx == 2:
        val = '=COUNTIF(\'📋 Task Tracker\'!J7:J106,"Complete")'
        apply_cell(ws_dash, kpi_vals_row, col, val, font=kpi_val_font, fill=kpi_card_fill,
                   alignment=center_align, border=kpi_card_border)
    elif idx == 3:
        val = '=SUM(\'💰 Budget & Resources\'!H7:H106)'
        apply_cell(ws_dash, kpi_vals_row, col, val, font=kpi_val_font, fill=kpi_card_fill,
                   alignment=center_align, border=kpi_card_border, number_format='$#,##0')
    elif idx == 4:
        val = '=SUM(\'💰 Budget & Resources\'!G7:G106)-SUM(\'💰 Budget & Resources\'!H7:H106)'
        apply_cell(ws_dash, kpi_vals_row, col, val, font=kpi_val_font, fill=kpi_card_fill,
                   alignment=center_align, border=kpi_card_border, number_format='$#,##0')
    elif idx == 5:
        val = '=MAX(0,NETWORKDAYS(TODAY(),MAX(\'📋 Task Tracker\'!E7:E106)))'
        apply_cell(ws_dash, kpi_vals_row, col, val, font=kpi_val_font, fill=kpi_card_fill,
                   alignment=center_align, border=kpi_card_border, number_format='0')
ws_dash.row_dimensions[kpi_vals_row].height = 48

# -- Project Info --
info_row = 8
section_header(ws_dash, info_row, 2, 7, "📋 PROJECT INFORMATION")
info_data = [
    ("Project Name:", "Website Redesign & Launch", "Project Manager:", "Zhang Wei"),
    ("Start Date:", "2025-06-01", "End Date:", "2025-12-31"),
    ("Department:", "Product & Engineering", "Priority:", "Critical"),
    ("Budget Total:", "$180,000", "Status:", "In Progress"),
]
r = info_row + 1
for row_data in info_data:
    for ci, (label, value) in enumerate([(0,1),(2,3)]):
        lbl_col = 2 + ci*3   # 2, 5
        val_col = 3 + ci*3   # 3, 6
        apply_cell(ws_dash, r, lbl_col, label, font=Font(name='Calibri', size=10, bold=True, color=NAVY),
                   fill=None, alignment=left_align, border=Border())
        apply_cell(ws_dash, r, val_col, value, font=data_font, fill=None, alignment=left_align, border=Border())
    r += 1

# Spacer
r += 1

# -- Task Status Breakdown --
section_header(ws_dash, r, 2, 7, "📈 TASK STATUS BREAKDOWN")
r += 1
status_labels = ["Not Started", "In Progress", "Complete", "Blocked", "Delayed"]
status_colors = [GRAY, STEEL, GREEN, RED, ORANGE]
status_fills = [
    PatternFill(start_color='EAECEE', end_color='EAECEE', fill_type='solid'),
    PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid'),
    PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid'),
    PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
    PatternFill(start_color='FDEBD0', end_color='FDEBD0', fill_type='solid'),
]
apply_cell(ws_dash, r, 2, "Status", font=Font(name='Calibri', size=10, bold=True, color=WHITE_BG), fill=hdr_fill, alignment=center_align, border=header_border)
apply_cell(ws_dash, r, 3, "Count", font=Font(name='Calibri', size=10, bold=True, color=WHITE_BG), fill=hdr_fill, alignment=center_align, border=header_border)
apply_cell(ws_dash, r, 4, "% of Total", font=Font(name='Calibri', size=10, bold=True, color=WHITE_BG), fill=hdr_fill, alignment=center_align, border=header_border)
ws_dash.row_dimensions[r].height = 28

for si, sl in enumerate(status_labels):
    r += 1
    count_formula = f'=COUNTIF(\'📋 Task Tracker\'!J7:J106,"{sl}")'
    pct_formula = f'=IF(SUM(B{r}:B{r+4})=0,0,B{r}/SUM(B{r-4}:B{r}))'
    apply_cell(ws_dash, r, 2, sl, font=data_font, fill=status_fills[si], alignment=left_align, border=thin_border)
    apply_cell(ws_dash, r, 3, count_formula, font=data_font, fill=status_fills[si], alignment=center_align, border=thin_border)
    apply_cell(ws_dash, r, 4, pct_formula, font=data_font, fill=status_fills[si], alignment=center_align, border=thin_border, number_format='0%')

total_row = r + 1
apply_cell(ws_dash, total_row, 2, "TOTAL", font=Font(name='Calibri', size=10, bold=True, color=NAVY), fill=None, alignment=left_align, border=thin_border)
apply_cell(ws_dash, total_row, 3, f'=SUM(B{total_row-5}:B{total_row-1})', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border)
apply_cell(ws_dash, total_row, 4, f'=SUM(C{total_row-5}:C{total_row-1})', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border, number_format='0%')

# -- Milestone Timeline --
r = total_row + 2
section_header(ws_dash, r, 2, 7, "🚩 KEY MILESTONES")
r += 1
milestone_headers = ["Milestone", "Target Date", "Status", "Owner", "Days Until"]
apply_header(ws_dash, r, milestone_headers, start_col=2)
r += 1
milestones = [
    ("Requirements Finalized", "2025-07-15", "Complete", "PM", 0),
    ("Design Prototype Approved", "2025-08-30", "Complete", "UX Lead", 0),
    ("Frontend Development Complete", "2025-10-15", "In Progress", "Tech Lead", 0),
    ("Backend Integration Done", "2025-11-15", "In Progress", "BE Lead", 0),
    ("QA & UAT Testing", "2025-12-01", "Not Started", "QA Lead", 0),
    ("Production Launch", "2025-12-31", "Not Started", "PM", 0),
]
for mi, m in enumerate(milestones):
    row = r + mi
    fill = alt_fill if row % 2 == 0 else None
    days_formula = f'=MAX(0,NETWORKDAYS(TODAY(),C{row}))'
    statuses = ["", "Not Started", "In Progress", "Complete", "Blocked", "Delayed"]
    apply_cell(ws_dash, row, 2, m[0], font=data_font, fill=fill, alignment=left_align, border=thin_border)
    apply_cell(ws_dash, row, 3, m[1], font=data_font, fill=fill, alignment=center_align, border=thin_border)
    # Status with dropdown
    apply_cell(ws_dash, row, 4, m[2], font=data_font, fill=fill, alignment=center_align, border=thin_border)
    apply_cell(ws_dash, row, 5, m[3], font=data_font, fill=fill, alignment=center_align, border=thin_border)
    apply_cell(ws_dash, row, 6, days_formula, font=data_font, fill=fill, alignment=center_align, border=thin_border, number_format='0')

# -- Risk Summary --
r = r + len(milestones) + 2
section_header(ws_dash, r, 2, 7, "⚠️ RISK OVERVIEW")
r += 1
risk_headers = ["Risk Level", "Count", "Mitigation Status"]
apply_header(ws_dash, r, risk_headers, start_col=2)
r += 1
for level in ["Critical", "High", "Medium", "Low"]:
    fill = alt_fill if r % 2 == 0 else None
    apply_cell(ws_dash, r, 2, level, font=data_font, fill=fill, alignment=left_align, border=thin_border)
    apply_cell(ws_dash, r, 3, f'=COUNTIF(\'🚩 Risk Register\'!G7:G56,"{level}")', font=data_font, fill=fill, alignment=center_align, border=thin_border)
    apply_cell(ws_dash, r, 4, "", font=data_font, fill=fill, alignment=center_align, border=thin_border)
    r += 1

# -- Conditional formatting: data bars on progress --
ws_dash.conditional_formatting.add(f'B{kpi_vals_row}',
    DataBarRule(start_type='min', end_type='max', color=STEEL, showValue=True))

# Data validation for milestone status
dv_status = DataValidation(type="list", formula1='"Not Started,In Progress,Complete,Blocked,Delayed"', allow_blank=True)
ws_dash.add_data_validation(dv_status)
# Add to milestone status cells
for mi in range(len(milestones)):
    dv_status.add(ws_dash.cell(row=14+mi, column=4))

print("  [OK] Dashboard sheet created")

# ─────────────────────────────────────────
# SHEET 2: TASK TRACKER (GANTT-STYLE)
# ─────────────────────────────────────────
ws_task = wb.create_sheet("📋 Task Tracker")
ws_task.sheet_properties.tabColor = STEEL

task_headers = [
    "WBS ID", "Task Name", "Description", "Department",
    "Start Date", "End Date", "Duration\n(Days)",
    "Predecessors", "Assigned To", "Status",
    "% Complete", "Actual Start", "Actual End",
    "Actual Days", "Variance\n(Days)", "Critical\nPath?", "Comments"
]
task_widths = [8, 22, 28, 14, 12, 12, 10, 12, 16, 13, 10, 12, 12, 10, 10, 10, 22]
set_col_widths(ws_task, task_widths)

merge_title(ws_task, 1, 1, 16, "📋 Project Task Tracker & Gantt Chart")
merge_subtitle(ws_task, 2, 1, 16, "Complete Work Breakdown Structure with dependencies, critical path & automatic duration calculations")
ws_task.row_dimensions[3].height = 6

# Header row
apply_header(ws_task, 4, task_headers)
ws_task.row_dimensions[4].height = 36

# Sample data - rich project tasks
tasks = [
    # WBS, Task, Desc, Dept, Start, End, Dur, Pred, Assigned, Status, %, ActStart, ActEnd, ActDays, Var, Critical, Comments
    ["1.0", "Project Initiation", "Kickoff, stakeholder alignment, project charter", "PMO", "2025-06-01", "2025-06-14", "", "", "Zhang Wei", "Complete", 100, "2025-06-01", "2025-06-13", "", "", "Yes", "Charter signed"],
    ["1.1", "Stakeholder Interviews", "Conduct 12 stakeholder discovery sessions", "PMO", "2025-06-01", "2025-06-07", "", "", "Li Na", "Complete", 100, "2025-06-01", "2025-06-06", "", "", "Yes", ""],
    ["1.2", "Project Charter Draft", "Write and circulate project charter for approval", "PMO", "2025-06-08", "2025-06-14", "", "1.1", "Zhang Wei", "Complete", 100, "2025-06-07", "2025-06-13", "", "", "Yes", ""],
    ["2.0", "Requirements Gathering", "User stories, functional & non-functional requirements", "Product", "2025-06-15", "2025-07-15", "", "1.0", "Chen Xiao", "Complete", 100, "2025-06-15", "2025-07-14", "", "", "Yes", ""],
    ["2.1", "User Research", "User interviews, surveys, competitive analysis", "UX", "2025-06-15", "2025-06-30", "", "1.2", "Wang Fang", "Complete", 100, "2025-06-15", "2025-07-01", "", "", "No", ""],
    ["2.2", "Functional Specs", "Document all functional requirements & user flows", "Product", "2025-07-01", "2025-07-10", "", "2.1", "Chen Xiao", "Complete", 100, "2025-07-01", "2025-07-09", "", "", "No", ""],
    ["2.3", "Technical Specs", "API design, data model, integration requirements", "Engineering", "2025-07-05", "2025-07-15", "", "2.1", "Zhao Lei", "Complete", 100, "2025-07-05", "2025-07-15", "", "", "No", ""],
    ["3.0", "Design Phase", "UI/UX design, prototyping & design system", "UX", "2025-07-16", "2025-09-15", "", "2.0", "Wang Fang", "In Progress", 85, "2025-07-16", "", "", "", "Yes", ""],
    ["3.1", "Wireframes", "Low-fidelity wireframes for all 24 screens", "UX", "2025-07-16", "2025-08-10", "", "2.3", "Wang Fang", "Complete", 100, "2025-07-16", "2025-08-09", "", "", "No", ""],
    ["3.2", "High-Fidelity Mockups", "Pixel-perfect designs for all screens", "UX", "2025-08-11", "2025-09-05", "", "3.1", "Liu Yang", "In Progress", 90, "2025-08-11", "", "", "", "No", ""],
    ["3.3", "Design System", "Component library, tokens, responsive breakpoints", "UX", "2025-08-15", "2025-09-15", "", "3.2", "Wang Fang", "In Progress", 70, "2025-08-15", "", "", "", "No", ""],
    ["3.4", "Prototype Testing", "Usability testing with 15 participants", "UX", "2025-09-01", "2025-09-14", "", "3.2", "Liu Yang", "Not Started", 0, "", "", "", "", "No", "Depends on mockup completion"],
    ["4.0", "Frontend Development", "React/Next.js implementation of all screens", "Engineering", "2025-09-01", "2025-11-15", "", "3.0", "Zhao Lei", "In Progress", 35, "2025-09-01", "", "", "", "Yes", ""],
    ["4.1", "Component Development", "Build all reusable UI components", "Engineering", "2025-09-01", "2025-10-15", "", "3.3", "Zhao Lei", "In Progress", 50, "2025-09-01", "", "", "", "No", ""],
    ["4.2", "Page Assembly", "Assemble pages, routing, state management", "Engineering", "2025-10-01", "2025-10-31", "", "4.1", "Sun Yue", "Not Started", 25, "2025-10-01", "", "", "", "No", ""],
    ["4.3", "API Integration", "Connect frontend to all backend endpoints", "Engineering", "2025-10-15", "2025-11-10", "", "4.1", "Zhao Lei", "Not Started", 0, "", "", "", "", "No", ""],
    ["4.4", "Performance Optimization", "Lighthouse score > 90, bundle optimization", "Engineering", "2025-11-01", "2025-11-15", "", "4.2,4.3", "Sun Yue", "Not Started", 0, "", "", "", "", "No", ""],
    ["5.0", "Backend Development", "API, database, auth, infrastructure", "Engineering", "2025-09-15", "2025-11-30", "", "2.0", "Huang Wei", "In Progress", 30, "2025-09-15", "", "", "", "Yes", ""],
    ["5.1", "Database Schema", "Design & implement PostgreSQL schema", "Engineering", "2025-09-15", "2025-09-30", "", "2.3", "Huang Wei", "Complete", 100, "2025-09-15", "2025-09-29", "", "", "No", ""],
    ["5.2", "API Development", "RESTful API endpoints, GraphQL layer", "Engineering", "2025-10-01", "2025-11-15", "", "5.1", "Huang Wei", "In Progress", 40, "2025-10-01", "", "", "", "No", ""],
    ["5.3", "Authentication System", "OAuth2.0, JWT, role-based access control", "Engineering", "2025-10-15", "2025-11-05", "", "5.1", "Ma Jing", "In Progress", 25, "2025-10-15", "", "", "", "No", ""],
    ["5.4", "CI/CD Pipeline", "GitHub Actions, Docker, AWS deployment", "DevOps", "2025-10-01", "2025-10-31", "", "5.1", "Zhou Tao", "In Progress", 60, "2025-10-01", "", "", "", "No", ""],
    ["6.0", "Quality Assurance", "Testing, bug fixing & UAT", "QA", "2025-11-16", "2025-12-15", "", "4.0,5.0", "Xu Jing", "Not Started", 0, "", "", "", "", "Yes", ""],
    ["6.1", "Test Plan Creation", "Test cases, test data, automation scripts", "QA", "2025-11-16", "2025-11-25", "", "4.4,5.4", "Xu Jing", "Not Started", 0, "", "", "", "", "No", ""],
    ["6.2", "Integration Testing", "End-to-end integration test execution", "QA", "2025-11-26", "2025-12-05", "", "6.1", "Xu Jing", "Not Started", 0, "", "", "", "", "No", ""],
    ["6.3", "UAT", "User acceptance testing with stakeholders", "QA", "2025-12-06", "2025-12-12", "", "6.2", "Zhang Wei", "Not Started", 0, "", "", "", "", "No", ""],
    ["6.4", "Bug Fixes", "Critical & high-priority bug resolution", "Engineering", "2025-12-01", "2025-12-15", "", "6.2", "Zhao Lei", "Not Started", 0, "", "", "", "", "No", ""],
    ["7.0", "Launch & Closeout", "Production deployment & project retrospective", "PMO", "2025-12-16", "2025-12-31", "", "6.0", "Zhang Wei", "Not Started", 0, "", "", "", "", "Yes", ""],
    ["7.1", "Production Deployment", "Go-live deployment & smoke testing", "DevOps", "2025-12-16", "2025-12-19", "", "6.3,6.4", "Zhou Tao", "Not Started", 0, "", "", "", "", "No", ""],
    ["7.2", "Post-Launch Monitoring", "Monitor KPIs, error rates, user feedback for 2 weeks", "Engineering", "2025-12-20", "2025-12-28", "", "7.1", "All Team", "Not Started", 0, "", "", "", "", "No", ""],
    ["7.3", "Retrospective", "Project retrospective & lessons learned documentation", "PMO", "2025-12-29", "2025-12-31", "", "7.2", "Zhang Wei", "Not Started", 0, "", "", "", "", "No", ""],
]

# Data validations
dv_status_task = DataValidation(type="list", formula1='"Not Started,In Progress,Complete,Blocked,Delayed"', allow_blank=True)
dv_dept = DataValidation(type="list", formula1='"PMO,Product,UX,Engineering,DevOps,QA,Marketing"', allow_blank=True)
dv_critical = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
ws_task.add_data_validation(dv_status_task)
ws_task.add_data_validation(dv_dept)
ws_task.add_data_validation(dv_critical)

data_start = 5
for i, t in enumerate(tasks):
    row = data_start + i
    row_fill = alt_fill if row % 2 == 0 else None

    # WBS ID
    apply_cell(ws_task, row, 1, t[0], font=Font(name='Consolas', size=9, color=DARK_TEXT), fill=row_fill, alignment=left_align, border=thin_border)
    # Task Name
    apply_cell(ws_task, row, 2, t[1], font=Font(name='Calibri', size=10, bold=True, color=DARK_TEXT), fill=row_fill, alignment=left_align, border=thin_border)
    # Description
    apply_cell(ws_task, row, 3, t[2], font=data_font, fill=row_fill, alignment=left_align, border=thin_border)
    # Department
    apply_cell(ws_task, row, 4, t[3], font=data_font, fill=row_fill, alignment=center_align, border=thin_border)
    dv_dept.add(ws_task.cell(row=row, column=4))
    # Start Date
    apply_cell(ws_task, row, 5, t[4], font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='YYYY-MM-DD')
    # End Date
    apply_cell(ws_task, row, 6, t[5], font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='YYYY-MM-DD')
    # Duration - NETWORKDAYS formula
    dur_formula = f'=IF(AND(E{row}<>"",F{row}<>""),NETWORKDAYS(E{row},F{row}),"")'
    apply_cell(ws_task, row, 7, dur_formula, font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='0')
    # Predecessors
    apply_cell(ws_task, row, 8, t[7], font=data_font, fill=row_fill, alignment=center_align, border=thin_border)
    # Assigned To
    apply_cell(ws_task, row, 9, t[8], font=data_font, fill=row_fill, alignment=left_align, border=thin_border)
    # Status
    apply_cell(ws_task, row, 10, t[9], font=data_font, fill=row_fill, alignment=center_align, border=thin_border)
    dv_status_task.add(ws_task.cell(row=row, column=10))
    # % Complete
    apply_cell(ws_task, row, 11, t[10]/100 if t[10] else 0, font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='0%')
    # Actual Start
    act_start = t[11] if t[11] else ""
    apply_cell(ws_task, row, 12, act_start, font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='YYYY-MM-DD')
    # Actual End
    act_end = t[12] if t[12] else ""
    apply_cell(ws_task, row, 13, act_end, font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='YYYY-MM-DD')
    # Actual Days
    act_days_formula = f'=IF(AND(L{row}<>"",M{row}<>""),NETWORKDAYS(L{row},M{row}),IF(AND(L{row}<>"",M{row}=""),NETWORKDAYS(L{row},TODAY()),""))'
    apply_cell(ws_task, row, 14, act_days_formula, font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='0')
    # Variance
    var_formula = f'=IF(AND(G{row}<>"",N{row}<>""),N{row}-G{row},"")'
    apply_cell(ws_task, row, 15, var_formula, font=data_font, fill=row_fill, alignment=center_align, border=thin_border, number_format='+0;-0;0')
    # Critical Path
    apply_cell(ws_task, row, 16, t[14], font=Font(name='Calibri', size=10, bold=True, color=RED if t[14]=='Yes' else DARK_TEXT), fill=row_fill, alignment=center_align, border=thin_border)
    dv_critical.add(ws_task.cell(row=row, column=16))
    # Comments
    apply_cell(ws_task, row, 17, t[15], font=data_font, fill=row_fill, alignment=left_align, border=thin_border)

last_task_row = data_start + len(tasks) - 1

# Summary row at bottom
summary_row = last_task_row + 3
section_header(ws_task, summary_row, 1, 17, "📊 TASK SUMMARY")
sr = summary_row + 1
apply_cell(ws_task, sr, 1, "", font=data_font, border=Border())
apply_cell(ws_task, sr, 2, "TOTAL TASKS:", font=Font(name='Calibri', size=10, bold=True, color=NAVY), alignment=right_align, border=Border())
apply_cell(ws_task, sr, 3, f'=COUNTA(B{data_start}:B{last_task_row})', font=Font(name='Calibri', size=14, bold=True, color=NAVY), alignment=center_align, border=Border())
apply_cell(ws_task, sr, 4, "COMPLETED:", font=Font(name='Calibri', size=10, bold=True, color=GREEN), alignment=right_align, border=Border())
apply_cell(ws_task, sr, 5, f'=COUNTIF(J{data_start}:J{last_task_row},"Complete")', font=Font(name='Calibri', size=14, bold=True, color=GREEN), alignment=center_align, border=Border())
apply_cell(ws_task, sr, 6, "IN PROGRESS:", font=Font(name='Calibri', size=10, bold=True, color=STEEL), alignment=right_align, border=Border())
apply_cell(ws_task, sr, 7, f'=COUNTIF(J{data_start}:J{last_task_row},"In Progress")', font=Font(name='Calibri', size=14, bold=True, color=STEEL), alignment=center_align, border=Border())
apply_cell(ws_task, sr, 8, "BLOCKED:", font=Font(name='Calibri', size=10, bold=True, color=RED), alignment=right_align, border=Border())
apply_cell(ws_task, sr, 9, f'=COUNTIF(J{data_start}:J{last_task_row},"Blocked")', font=Font(name='Calibri', size=14, bold=True, color=RED), alignment=center_align, border=Border())
apply_cell(ws_task, sr, 10, "AVG PROGRESS:", font=Font(name='Calibri', size=10, bold=True, color=PURPLE), alignment=right_align, border=Border())
apply_cell(ws_task, sr, 11, f'=AVERAGE(K{data_start}:K{last_task_row})', font=Font(name='Calibri', size=14, bold=True, color=PURPLE), alignment=center_align, border=Border(), number_format='0%')

# Conditional formatting
# Data bars on % Complete
ws_task.conditional_formatting.add(f'K{data_start}:K{last_task_row}',
    DataBarRule(start_type='min', end_type='max', color=STEEL, showValue=True))

# Status color rules
for status, color in [("Complete", 'D5F5E3'), ("In Progress", 'D6EAF8'), ("Blocked", 'FADBD8'),
                        ("Delayed", 'FDEBD0'), ("Not Started", 'EAECEE')]:
    ws_task.conditional_formatting.add(f'J{data_start}:J{last_task_row}',
        CellIsRule(operator='equal', formula=[f'"{status}"'], fill=PatternFill(start_color=color, end_color=color, fill_type='solid')))

# Variance > 0 (over schedule) in red
ws_task.conditional_formatting.add(f'O{data_start}:O{last_task_row}',
    CellIsRule(operator='greaterThan', formula=['0'], font=Font(name='Calibri', size=10, color='C0392B', bold=True),
               fill=PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')))

# Highlight critical path
ws_task.conditional_formatting.add(f'A{data_start}:Q{last_task_row}',
    FormulaRule(formula=[f'$P{data_start}="Yes"'], font=Font(name='Calibri', size=10, bold=True, color=RED),
                fill=PatternFill(start_color='FDEDEC', end_color='FDEDEC', fill_type='solid')))

# Frozen panes
ws_task.freeze_panes = 'A5'

print("  [OK] Task Tracker sheet created")

# ─────────────────────────────────────────
# SHEET 3: BUDGET & RESOURCES
# ─────────────────────────────────────────
ws_budget = wb.create_sheet("💰 Budget & Resources")
ws_budget.sheet_properties.tabColor = GREEN

budget_headers = [
    "Resource ID", "Resource Name", "Type", "Department",
    "Rate\n($/hr or $/day)", "Unit", "Allocated\nHours/Days",
    "Budget ($)", "Actual ($)", "Variance ($)",
    "Variance %", "Utilization %", "Assigned Tasks", "Notes"
]
budget_widths = [10, 20, 12, 14, 12, 8, 12, 14, 14, 14, 12, 12, 28, 18]
set_col_widths(ws_budget, budget_widths)

merge_title(ws_budget, 1, 1, 14, "💰 Budget & Resource Allocation")
merge_subtitle(ws_budget, 2, 1, 14, "Track all project resources, budget allocation, actual spend & cost variance analysis")
ws_budget.row_dimensions[3].height = 6
apply_header(ws_budget, 4, budget_headers)

resources = [
    # ID, Name, Type, Dept, Rate, Unit, Alloc, Budget, Actual, Var, Var%, Util%, Tasks, Notes
    ["RES-01", "Zhang Wei", "Human", "PMO", 75, "hr", 480, "", "", "", "", "", "1.2,1.3,7.3", "Senior PM"],
    ["RES-02", "Li Na", "Human", "PMO", 55, "hr", 320, "", "", "", "", "", "1.1", ""],
    ["RES-03", "Wang Fang", "Human", "UX", 65, "hr", 520, "", "", "", "", "", "2.1,3.1,3.3", "UX Lead"],
    ["RES-04", "Liu Yang", "Human", "UX", 50, "hr", 480, "", "", "", "", "", "3.2,3.4", ""],
    ["RES-05", "Zhao Lei", "Human", "Engineering", 80, "hr", 640, "", "", "", "", "", "2.3,4.1,4.3,6.4", "Tech Lead"],
    ["RES-06", "Sun Yue", "Human", "Engineering", 60, "hr", 480, "", "", "", "", "", "4.2,4.4", ""],
    ["RES-07", "Huang Wei", "Human", "Engineering", 70, "hr", 560, "", "", "", "", "", "5.1,5.2", "BE Lead"],
    ["RES-08", "Ma Jing", "Human", "Engineering", 55, "hr", 400, "", "", "", "", "", "5.3", ""],
    ["RES-09", "Zhou Tao", "Human", "DevOps", 75, "hr", 360, "", "", "", "", "", "5.4,7.1", "DevOps Lead"],
    ["RES-10", "Xu Jing", "Human", "QA", 60, "hr", 400, "", "", "", "", "", "6.1,6.2,6.3", "QA Lead"],
    ["RES-11", "Chen Xiao", "Human", "Product", 70, "hr", 320, "", "", "", "", "", "2.0,2.2", "Product Mgr"],
    ["RES-12", "AWS Cloud", "Infrastructure", "DevOps", 5000, "mo", 6, "", "", "", "", "", "", "$5k/month for 6 months"],
    ["RES-13", "Figma Enterprise", "Software", "UX", 75, "mo", 7, "", "", "", "", "", "", "Per seat/month × 3 seats"],
    ["RES-14", "GitHub Enterprise", "Software", "Engineering", 42, "mo", 7, "", "", "", "", "", "", "Per seat/month × 8 seats"],
    ["RES-15", "Testing Devices", "Equipment", "QA", 0, "one", 1, 3000, "", "", "", "", "", "Mobile + tablet test lab"],
    ["RES-16", "SSL Certificate", "Software", "DevOps", 0, "one", 1, 1500, "", "", "", "", "", "Wildcard cert"],
    ["RES-17", "Analytics Tool", "Software", "Product", 199, "mo", 6, "", "", "", "", "", "", "Mixpanel/Amplitude"],
    ["RES-18", "Training Budget", "Service", "PMO", 0, "one", 1, 5000, "", "", "", "", "", "Team upskilling"],
    ["RES-19", "Contingency", "Reserve", "PMO", 0, "one", 1, 18000, "", "", "", "", "", "10% contingency"],
    ["RES-20", "Content Writer", "Human", "Marketing", 45, "hr", 160, "", "", "", "", "", "", "Launch content"],
]

dv_resource_type = DataValidation(type="list", formula1='"Human,Infrastructure,Software,Equipment,Service,Reserve"', allow_blank=True)
dv_department = DataValidation(type="list", formula1='"PMO,Product,UX,Engineering,DevOps,QA,Marketing"', allow_blank=True)
ws_budget.add_data_validation(dv_resource_type)
ws_budget.add_data_validation(dv_department)

budget_start = 5
for i, res in enumerate(resources):
    row = budget_start + i
    rf = alt_fill if row % 2 == 0 else None

    apply_cell(ws_budget, row, 1, res[0], font=Font(name='Consolas', size=9), fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_budget, row, 2, res[1], font=Font(name='Calibri', size=10, bold=True), fill=rf, alignment=left_align, border=thin_border)
    apply_cell(ws_budget, row, 3, res[2], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    dv_resource_type.add(ws_budget.cell(row=row, column=3))
    apply_cell(ws_budget, row, 4, res[3], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    dv_department.add(ws_budget.cell(row=row, column=4))
    apply_cell(ws_budget, row, 5, res[4], font=data_font, fill=rf, alignment=center_align, border=thin_border,
               number_format='$#,##0' if res[4] > 100 else '$#,##0.00')
    apply_cell(ws_budget, row, 6, res[5], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_budget, row, 7, res[6], font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='#,##0')

    # Budget formula: Rate * Allocated (for hourly resources) or flat amount for others
    if res[2] == "Human":
        budget_formula = f'=E{row}*G{row}'
    elif res[5] == "mo":
        budget_formula = f'=E{row}*G{row}'
    else:
        budget_formula = res[7] if res[7] else 0
    apply_cell(ws_budget, row, 8, budget_formula, font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')

    # Actual - some sample actuals
    actual_val = res[8] if res[8] else ""
    if actual_val == "" and res[2] == "Human":
        actual_formula = f'=IF(OR(C{row}="Reserve"),0,E{row}*G{row}*RAND()*1.1)'
        apply_cell(ws_budget, row, 9, actual_formula, font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    else:
        apply_cell(ws_budget, row, 9, actual_val, font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')

    # Variance
    apply_cell(ws_budget, row, 10, f'=IF(H{row}="","",H{row}-I{row})', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    # Variance %
    apply_cell(ws_budget, row, 11, f'=IF(H{row}=0,"",(H{row}-I{row})/H{row})', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='0.0%')
    # Utilization
    apply_cell(ws_budget, row, 12, f'=IF(OR(C{row}="Human",C{row}="Software"),MIN(1,I{row}/H{row}),"")', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='0%')
    # Tasks
    apply_cell(ws_budget, row, 13, res[12], font=Font(name='Calibri', size=8, color=MED_TEXT), fill=rf, alignment=left_align, border=thin_border)
    # Notes
    apply_cell(ws_budget, row, 14, res[13], font=Font(name='Calibri', size=8, color=MED_TEXT), fill=rf, alignment=left_align, border=thin_border)

budget_last = budget_start + len(resources) - 1

# Budget totals
br = budget_last + 3
section_header(ws_budget, br, 1, 14, "📊 BUDGET SUMMARY")
br += 1
apply_cell(ws_budget, br, 1, "", font=data_font, border=Border())
apply_cell(ws_budget, br, 2, "TOTAL BUDGET:", font=Font(name='Calibri', size=12, bold=True, color=NAVY), fill=None, alignment=right_align, border=Border())
apply_cell(ws_budget, br, 3, f'=SUM(H{budget_start}:H{budget_last})', font=Font(name='Calibri', size=16, bold=True, color=NAVY), fill=None, alignment=center_align, border=Border(), number_format='$#,##0')
apply_cell(ws_budget, br, 4, "TOTAL ACTUAL:", font=Font(name='Calibri', size=12, bold=True, color=DARK_BLUE), fill=None, alignment=right_align, border=Border())
apply_cell(ws_budget, br, 5, f'=SUM(I{budget_start}:I{budget_last})', font=Font(name='Calibri', size=16, bold=True, color=DARK_BLUE), fill=None, alignment=center_align, border=Border(), number_format='$#,##0')
apply_cell(ws_budget, br, 6, "VARIANCE:", font=Font(name='Calibri', size=12, bold=True, color=PURPLE), fill=None, alignment=right_align, border=Border())
apply_cell(ws_budget, br, 7, f'=B{br}-D{br}', font=Font(name='Calibri', size=16, bold=True, color=PURPLE), fill=None, alignment=center_align, border=Border(), number_format='$#,##0')
apply_cell(ws_budget, br, 8, "CONTINGENCY REMAINING:", font=Font(name='Calibri', size=10, bold=True, color=ORANGE), fill=None, alignment=right_align, border=Border())
apply_cell(ws_budget, br, 9, f'=H{budget_start+17}-I{budget_start+17}', font=Font(name='Calibri', size=12, bold=True, color=ORANGE), fill=None, alignment=center_align, border=Border(), number_format='$#,##0')

# Department breakdown
br2 = br + 2
section_header(ws_budget, br2, 1, 14, "📊 BUDGET BY DEPARTMENT")
br2 += 1
apply_header(ws_budget, br2, ["Department", "Budget ($)", "Actual ($)", "Variance ($)", "Variance %"], start_col=2)
depts = [("PMO", 2), ("Product", 4), ("UX", 5), ("Engineering", 8), ("DevOps", 10), ("QA", 11), ("Marketing", 14)]
for di, (dept, ref_col) in enumerate(depts):
    row = br2 + 1 + di
    rf = alt_fill if row % 2 == 0 else None
    apply_cell(ws_budget, row, 2, dept, font=Font(name='Calibri', size=10, bold=True), fill=rf, alignment=left_align, border=thin_border)
    apply_cell(ws_budget, row, 3, f'=SUMIF(D{budget_start}:D{budget_last},B{row},H{budget_start}:H{budget_last})', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    apply_cell(ws_budget, row, 4, f'=SUMIF(D{budget_start}:D{budget_last},B{row},I{budget_start}:I{budget_last})', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    apply_cell(ws_budget, row, 5, f'=C{row}-D{row}', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    apply_cell(ws_budget, row, 6, f'=IF(C{row}=0,"",E{row}/C{row})', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='0.0%')

# Conditional formatting
ws_budget.conditional_formatting.add(f'J{budget_start}:J{budget_last}',
    CellIsRule(operator='lessThan', formula=['0'], font=Font(color='C0392B', bold=True),
               fill=PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')))
ws_budget.conditional_formatting.add(f'J{budget_start}:J{budget_last}',
    CellIsRule(operator='greaterThan', formula=['0'], font=Font(color='1B7A3D', bold=True),
               fill=PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')))

print("  [OK] Budget & Resources sheet created")

# ─────────────────────────────────────────
# SHEET 4: RISK REGISTER
# ─────────────────────────────────────────
ws_risk = wb.create_sheet("🚩 Risk Register")
ws_risk.sheet_properties.tabColor = RED

risk_headers = [
    "Risk ID", "Risk Description", "Category", "Probability",
    "Impact", "Risk Score", "Risk Level", "Mitigation Strategy",
    "Owner", "Status", "Date Identified", "Last Reviewed"
]
risk_widths = [8, 32, 14, 12, 12, 12, 12, 32, 14, 14, 14, 14]
set_col_widths(ws_risk, risk_widths)

merge_title(ws_risk, 1, 1, 12, "🚩 Project Risk Register")
merge_subtitle(ws_risk, 2, 1, 12, "Identify, assess & track project risks with probability-impact matrix and mitigation planning")
ws_risk.row_dimensions[3].height = 6
apply_header(ws_risk, 4, risk_headers)

risks = [
    ["R-01", "Key developer departure mid-project", "Resource", "Medium", "Critical", "", "", "Cross-train developers, maintain thorough documentation, identify backup resources", "PM", "Monitored", "2025-06-01", "2025-09-15"],
    ["R-02", "Third-party API rate limit changes", "Technical", "Medium", "High", "", "", "Implement caching layer, negotiate enterprise SLA with vendor, build fallback queue", "Tech Lead", "Active", "2025-06-15", "2025-10-01"],
    ["R-03", "Scope creep from stakeholder requests", "Scope", "High", "High", "", "", "Formal change control process, weekly stakeholder reviews, prioritized backlog", "PM", "Active", "2025-06-01", "2025-09-20"],
    ["R-04", "Design approval delays from client", "Schedule", "High", "Medium", "", "", "Set clear approval SLAs in contract, schedule buffer days, parallel review tracks", "PM", "Monitored", "2025-07-01", "2025-10-15"],
    ["R-05", "Database performance under load", "Technical", "Medium", "High", "", "", "Load testing from sprint 1, query optimization, read replicas, caching strategy", "BE Lead", "Active", "2025-07-15", "2025-11-01"],
    ["R-06", "Browser compatibility issues", "Technical", "Low", "Medium", "", "", "Browserstack testing, progressive enhancement, polyfill strategy, supported browser matrix", "QA Lead", "Monitored", "2025-08-01", "2025-11-15"],
    ["R-07", "Budget overrun on cloud infrastructure", "Financial", "Medium", "Medium", "", "", "AWS budget alerts, auto-scaling limits, reserved instances, monthly cost review", "DevOps", "Active", "2025-06-01", "2025-10-01"],
    ["R-08", "Security vulnerability discovered late", "Security", "Low", "Critical", "", "", "SAST/DAST in CI pipeline, quarterly pen testing, dependency scanning, OWASP checklist", "Tech Lead", "Monitored", "2025-06-01", "2025-11-01"],
    ["R-09", "UAT reveals major UX issues", "Quality", "Medium", "High", "", "", "Early prototype testing with real users, iterative design, usability benchmarks", "UX Lead", "Monitored", "2025-08-15", "2025-11-20"],
    ["R-10", "Launch date conflict with marketing event", "Schedule", "Low", "Low", "", "", "Coordinate with marketing team quarterly, maintain flexible launch window options", "PM", "Monitored", "2025-09-01", "2025-12-01"],
    ["R-11", "Data migration failure", "Technical", "Low", "Critical", "", "", "Dry-run migrations, rollback plan, data validation checks, backup before each migration", "BE Lead", "Active", "2025-10-01", "2025-11-15"],
    ["R-12", "Vendor lock-in with proprietary service", "Strategic", "Low", "Medium", "", "", "Abstract vendor interfaces, evaluate open-source alternatives, maintain exit strategy", "Tech Lead", "Monitored", "2025-06-15", "2025-10-15"],
]

dv_probability = DataValidation(type="list", formula1='"Low,Medium,High,Critical"', allow_blank=True)
dv_impact = DataValidation(type="list", formula1='"Low,Medium,High,Critical"', allow_blank=True)
dv_risk_status = DataValidation(type="list", formula1='"Active,Monitored,Resolved,Closed"', allow_blank=True)
dv_risk_cat = DataValidation(type="list", formula1='"Technical,Resource,Scope,Schedule,Financial,Security,Quality,Strategic"', allow_blank=True)
ws_risk.add_data_validation(dv_probability)
ws_risk.add_data_validation(dv_impact)
ws_risk.add_data_validation(dv_risk_status)
ws_risk.add_data_validation(dv_risk_cat)

risk_start = 5
for i, rk in enumerate(risks):
    row = risk_start + i
    rf = alt_fill if row % 2 == 0 else None

    apply_cell(ws_risk, row, 1, rk[0], font=Font(name='Consolas', size=9), fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_risk, row, 2, rk[1], font=data_font, fill=rf, alignment=left_align, border=thin_border)
    apply_cell(ws_risk, row, 3, rk[2], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    dv_risk_cat.add(ws_risk.cell(row=row, column=3))
    apply_cell(ws_risk, row, 4, rk[3], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    dv_probability.add(ws_risk.cell(row=row, column=4))
    apply_cell(ws_risk, row, 5, rk[4], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    dv_impact.add(ws_risk.cell(row=row, column=5))
    # Risk Score formula: Probability × Impact matrix
    risk_score_formula = f'=IF(OR(D{row}="",E{row}=""),"",MATCH(D{row},{{"Low","Medium","High","Critical"}},0)*MATCH(E{row},{{"Low","Medium","High","Critical"}},0))'
    apply_cell(ws_risk, row, 6, risk_score_formula, font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='0')
    # Risk Level formula
    risk_level_formula = f'=IF(F{row}="","",IF(F{row}>=12,"Critical",IF(F{row}>=8,"High",IF(F{row}>=4,"Medium","Low"))))'
    apply_cell(ws_risk, row, 7, risk_level_formula, font=Font(name='Calibri', size=10, bold=True), fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_risk, row, 8, rk[7], font=data_font, fill=rf, alignment=left_align, border=thin_border)
    apply_cell(ws_risk, row, 9, rk[8], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_risk, row, 10, rk[9], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    dv_risk_status.add(ws_risk.cell(row=row, column=10))
    apply_cell(ws_risk, row, 11, rk[10], font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='YYYY-MM-DD')
    apply_cell(ws_risk, row, 12, rk[11], font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='YYYY-MM-DD')

risk_last = risk_start + len(risks) - 1

# Risk summary
rr = risk_last + 2
section_header(ws_risk, rr, 1, 12, "📊 RISK SUMMARY BY LEVEL")
rr += 1
apply_header(ws_risk, rr, ["Risk Level", "Count", "% of Total"], start_col=2)
for li, level in enumerate(["Critical", "High", "Medium", "Low"]):
    row = rr + 1 + li
    rf = alt_fill if row % 2 == 0 else None
    apply_cell(ws_risk, row, 2, level, font=Font(name='Calibri', size=10, bold=True, color=RED if level in ('Critical','High') else DARK_TEXT), fill=rf, alignment=left_align, border=thin_border)
    apply_cell(ws_risk, row, 3, f'=COUNTIF(G{risk_start}:G{risk_last},B{row})', font=data_font, fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_risk, row, 4, f'=IF(COUNTA(G{risk_start}:G{risk_last})=0,0,C{row}/COUNTA(G{risk_start}:G{risk_last}))', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='0%')

# Conditional formatting for risk score
ws_risk.conditional_formatting.add(f'F{risk_start}:F{risk_last}',
    ColorScaleRule(start_type='min', start_color='1B7A3D', mid_type='percentile', mid_value=50, mid_color='F1C40F',
                   end_type='max', end_color='C0392B'))

print("  [OK] Risk Register sheet created")

# ─────────────────────────────────────────
# SHEET 5: MONTHLY PROGRESS TRACKER
# ─────────────────────────────────────────
ws_monthly = wb.create_sheet("📈 Monthly Progress")
ws_monthly.sheet_properties.tabColor = PURPLE

monthly_headers = [
    "Month", "Planned Tasks", "Completed Tasks", "Completion Rate",
    "Planned Budget ($)", "Actual Spend ($)", "Budget Variance ($)",
    "Risks Open", "Risks Closed", "Team Velocity", "Cumulative Progress %"
]
monthly_widths = [12, 14, 14, 14, 16, 16, 16, 12, 12, 14, 18]
set_col_widths(ws_monthly, monthly_widths)

merge_title(ws_monthly, 1, 1, 11, "📈 Monthly Progress & Velocity Tracker")
merge_subtitle(ws_monthly, 2, 1, 11, "Track month-over-month project progress, team velocity, budget burn & risk trends")
ws_monthly.row_dimensions[3].height = 6
apply_header(ws_monthly, 4, monthly_headers)

months_data = [
    ["2025-06", 8, 8, "", 18000, 16500, "", 5, 0, 32, ""],
    ["2025-07", 10, 10, "", 25000, 23200, "", 7, 2, 38, ""],
    ["2025-08", 12, 11, "", 28000, 27100, "", 9, 3, 35, ""],
    ["2025-09", 14, 12, "", 32000, 30100, "", 10, 2, 42, ""],
    ["2025-10", 16, 8, "", 35000, 0, "", 11, 1, 0, ""],
    ["2025-11", 14, 0, "", 30000, 0, "", 8, 0, 0, ""],
    ["2025-12", 10, 0, "", 22000, 0, "", 5, 0, 0, ""],
]

monthly_start = 5
cum_prog = 0
for i, md in enumerate(months_data):
    row = monthly_start + i
    rf = alt_fill if row % 2 == 0 else None

    apply_cell(ws_monthly, row, 1, md[0], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_monthly, row, 2, md[1], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_monthly, row, 3, md[2], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    # Completion rate
    apply_cell(ws_monthly, row, 4, f'=IF(B{row}=0,"",C{row}/B{row})', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='0%')
    apply_cell(ws_monthly, row, 5, md[4], font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    apply_cell(ws_monthly, row, 6, md[5] if md[5] else "", font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    # Budget variance
    apply_cell(ws_monthly, row, 7, f'=IF(AND(E{row}<>"",F{row}<>""),E{row}-F{row},"")', font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='$#,##0')
    apply_cell(ws_monthly, row, 8, md[7], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_monthly, row, 9, md[8], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    apply_cell(ws_monthly, row, 10, md[9], font=data_font, fill=rf, alignment=center_align, border=thin_border)
    # Cumulative progress
    if i == 0:
        cum_formula = f'=C{row}/SUM($B${monthly_start}:$B${monthly_start+len(months_data)-1})'
    else:
        cum_formula = f'=SUM($C${monthly_start}:C{row})/SUM($B${monthly_start}:$B${monthly_start+len(months_data)-1})'
    apply_cell(ws_monthly, row, 11, cum_formula, font=data_font, fill=rf, alignment=center_align, border=thin_border, number_format='0%')

monthly_last = monthly_start + len(months_data) - 1

# Totals row
mr = monthly_last + 2
apply_cell(ws_monthly, mr, 1, "TOTAL", font=Font(name='Calibri', size=10, bold=True, color=NAVY), alignment=left_align, border=thin_border)
apply_cell(ws_monthly, mr, 2, f'=SUM(B{monthly_start}:B{monthly_last})', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border)
apply_cell(ws_monthly, mr, 3, f'=SUM(C{monthly_start}:C{monthly_last})', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border)
apply_cell(ws_monthly, mr, 4, f'=IF(B{mr}=0,"",C{mr}/B{mr})', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border, number_format='0%')
apply_cell(ws_monthly, mr, 5, f'=SUM(E{monthly_start}:E{monthly_last})', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border, number_format='$#,##0')
apply_cell(ws_monthly, mr, 6, f'=SUM(F{monthly_start}:F{monthly_last})', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border, number_format='$#,##0')
apply_cell(ws_monthly, mr, 7, f'=E{mr}-F{mr}', font=Font(name='Calibri', size=10, bold=True), alignment=center_align, border=thin_border, number_format='$#,##0')

# Conditional formatting for budget variance
ws_monthly.conditional_formatting.add(f'G{monthly_start}:G{monthly_last}',
    ColorScaleRule(start_type='min', start_color='C0392B', end_type='max', end_color='1B7A3D'))

# Add a line chart for cumulative progress
chart = LineChart()
chart.title = "Cumulative Project Progress"
chart.style = 10
chart.y_axis.title = "Progress %"
chart.x_axis.title = "Month"
chart.height = 15; chart.width = 22
data_ref = Reference(ws_monthly, min_col=11, min_row=monthly_start, max_row=monthly_last)
cats_ref = Reference(ws_monthly, min_col=1, min_row=monthly_start, max_row=monthly_last)
chart.add_data(data_ref, titles_from_data=False)
chart.set_categories(cats_ref)
chart.series[0].name = "Cumulative Progress"
chart.series[0].graphicalProperties.solidFill = STEEL
ws_monthly.add_chart(chart, f"A{mr + 3}")

print("  [OK] Monthly Progress sheet created")

# ─────────────────────────────────────────
# SHEET 6: INSTRUCTIONS
# ─────────────────────────────────────────
ws_ins = wb.create_sheet("📖 Instructions")
ws_ins.sheet_properties.tabColor = NAVY
set_col_widths(ws_ins, [5, 95])

def mr(ws, r, v, font, h=28):
    ws.merge_cells(f'A{r}:B{r}')
    ws.cell(row=r, column=1, value=v).font = font
    ws.cell(row=r, column=1).alignment = left_align
    ws.row_dimensions[r].height = h

mr(ws_ins, 1, "How to Use: Project Timeline & Gantt Chart (Professional Edition)", Font(name='Calibri', size=18, bold=True, color=NAVY), 38)
mr(ws_ins, 3, "Overview", Font(name='Calibri', size=13, bold=True, color=DARK_BLUE), 24)
mr(ws_ins, 4, "This is a comprehensive multi-dimensional project management template with 5 interconnected sheets. "
               "It combines task tracking (WBS + Gantt), resource budgeting, risk management, and monthly progress analytics — "
               "all linked through formulas for real-time project intelligence. Every key metric flows automatically to the Dashboard.", data_font, 48)

instructions_steps = [
    ("Project Setup", [
        "Go to 📊 Dashboard and fill in your project name, dates, budget total, and manager info.",
        "Review the pre-populated sample data to understand the template structure.",
        "Adjust the sample milestones to match your project timeline."
    ]),
    ("Task Planning (WBS)", [
        "Navigate to 📋 Task Tracker — this is your Work Breakdown Structure (WBS).",
        "Enter each task with a WBS ID, name, description, and department.",
        "Set Start Date and End Date — the Duration column auto-calculates using NETWORKDAYS().",
        "Define Predecessors by referencing other WBS IDs to build dependency chains.",
        "Mark tasks on the Critical Path as 'Yes' to highlight them in red across the row.",
        "Use the dropdown menus for Status and Department to maintain data consistency."
    ]),
    ("Resource & Budget Management", [
        "Go to 💰 Budget & Resources to plan all project costs.",
        "Enter each human resource with their hourly rate and allocated hours — budget auto-calculates.",
        "Add infrastructure, software licenses, equipment, and services.",
        "Track actual spend vs budget — variance columns highlight over/under automatically.",
        "The contingency reserve (10%) is built-in; monitor it on the budget summary row.",
        "Department budget breakdown uses SUMIF to aggregate costs by team."
    ]),
    ("Risk Management", [
        "Open 🚩 Risk Register to identify and assess project risks.",
        "For each risk, select Probability and Impact from dropdowns.",
        "The Risk Score auto-calculates (P×I matrix) and Risk Level is derived automatically.",
        "Use the conditional formatting (green→yellow→red) to prioritize mitigation efforts.",
        "Update Status as risks evolve: Active → Monitored → Resolved → Closed.",
        "Review risk trends on the 📈 Monthly Progress sheet alongside budget data."
    ]),
    ("Progress Tracking", [
        "📈 Monthly Progress captures month-over-month velocity and budget burn.",
        "Enter actual completed tasks and spend each month.",
        "The Cumulative Progress % column tracks overall project trajectory.",
        "Use the built-in line chart to visualize progress toward your targets.",
        "Compare planned vs actual metrics to identify schedule or budget slippage early."
    ]),
    ("Real-Time Dashboard", [
        "📊 Dashboard aggregates data from ALL other sheets automatically.",
        "KPI cards show Overall Progress (%), Total Tasks, Completed, Budget metrics, and Days Remaining.",
        "Task Status Breakdown uses COUNTIF across the Task Tracker.",
        "Milestones section shows upcoming dates and auto-calculates Days Until.",
        "Risk Overview summarizes risk levels from the Risk Register."
    ]),
]

r = 6
for title, steps in instructions_steps:
    mr(ws_ins, r, title, Font(name='Calibri', size=12, bold=True, color=STEEL), 26)
    r += 1
    for step in steps:
        mr(ws_ins, r, f"  {step}", data_font, 22)
        r += 1
    r += 1

mr(ws_ins, r, "Key Formulas Reference", Font(name='Calibri', size=12, bold=True, color=STEEL), 26)
formulas_ref = [
    "NETWORKDAYS(start, end) — Calculates working days between two dates (excludes weekends)",
    "COUNTIF(range, criteria) — Counts cells matching a condition (e.g., status = 'Complete')",
    "SUMIF(range, criteria, sum_range) — Sums values matching a department or category",
    "IF(condition, true_val, false_val) — Conditional logic for variance, status, levels",
    "MATCH + INDEX — Used in Risk Score matrix (Probability × Impact)",
    "Data Validation dropdowns — Ensure data consistency with predefined options",
    "Conditional Formatting — Auto-color cells based on values (data bars, color scales, rules)",
    "SUM / AVERAGE / MAX / MIN — Standard aggregation across all sheets",
]
r += 1
for f_ref in formulas_ref:
    mr(ws_ins, r, f"  = {f_ref}", Font(name='Consolas', size=9, color='555555'), 19)
    r += 1

r += 1
mr(ws_ins, r, "Pro Tips for Professional Project Management", Font(name='Calibri', size=12, bold=True, color=STEEL), 26)
r += 1
tips = [
    "💡 Update the Task Tracker weekly and review the Dashboard on Monday mornings — 15 minutes saves hours of confusion.",
    "💡 Use the Critical Path column to identify tasks that cannot slip without delaying the entire project.",
    "💡 Set calendar reminders to update Actual Spend in Budget & Resources at least bi-weekly.",
    "💡 For the Risk Register, review and update risk status during every sprint retrospective.",
    "💡 Export the Monthly Progress chart to include in stakeholder status reports.",
    "💡 Customize the dropdown lists (Data → Data Validation) to match your organization's departments and categories.",
    "💡 Protect sheets (Review → Protect Sheet) once your project data is finalized to prevent accidental edits.",
    "💡 Use Filter (Ctrl+Shift+L) on the Task Tracker to focus on specific departments, statuses, or assignees.",
]
for tip in tips:
    mr(ws_ins, r, tip, Font(name='Calibri', size=10, color='333333'), 24)
    r += 1

print("  [OK] Instructions sheet created")

# ─────────────────────────────────────────
# SHEET 7: ABOUT
# ─────────────────────────────────────────
ws_abt = wb.create_sheet("ℹ️ About")
ws_abt.sheet_properties.tabColor = '1F4E79'
set_col_widths(ws_abt, [5, 95])

mr(ws_abt, 1, "About: Project Timeline & Gantt Chart (Professional Edition)", Font(name='Calibri', size=18, bold=True, color=NAVY), 38)
about_items = [
    ("Template Name", "Project Timeline & Gantt Chart — Professional Edition"),
    ("Category", "Business & Management / Project Management"),
    ("Version", "2.0 — Multi-Sheet Professional Edition"),
    ("Last Updated", TD),
    ("Complexity Level", "Advanced — 5 interconnected data sheets + Dashboard + auto-calculations"),
    ("Formula Count", "200+ built-in formulas (NETWORKDAYS, COUNTIF, SUMIF, IF, MATCH, nested logic)"),
    ("Interactive Elements", "Data validation dropdowns, conditional formatting, data bars, color scales, charts"),
    ("Compatibility", "Microsoft Excel 2010+, Google Sheets, LibreOffice Calc (minor formatting variations)"),
    ("License", "MIT — Free for personal & commercial use"),
    ("Website", SITE),
]
for i, (k, v) in enumerate(about_items):
    row = 3 + i
    ws_abt.cell(row=row, column=1, value=f"{k}:").font = Font(name='Calibri', size=10, bold=True, color=NAVY)
    ws_abt.cell(row=row, column=2, value=v).font = data_font if k != "Website" else link_font
    ws_abt.row_dimensions[row].height = 22
    if k == "Website":
        ws_abt.cell(row=row, column=2).hyperlink = SITE

r = 3 + len(about_items) + 1
mr(ws_abt, r, "Sheet Structure", Font(name='Calibri', size=13, bold=True, color=DARK_BLUE), 24)
r += 1
sheets_desc = [
    "📊 Dashboard — Real-time KPIs, status breakdown, milestones, risk overview",
    "📋 Task Tracker — 30-task WBS with Gantt data, dependencies, critical path, auto-duration",
    "💰 Budget & Resources — 20 resources across departments with budget vs actual and variance",
    "🚩 Risk Register — 12 risks with probability-impact matrix, scoring, and mitigation",
    "📈 Monthly Progress — 7-month velocity tracker with cumulative progress chart",
    "📖 Instructions — Complete step-by-step tutorial and formula reference",
    "ℹ️ About — This sheet",
]
for sd in sheets_desc:
    mr(ws_abt, r, f"  {sd}", data_font, 21)
    r += 1

r += 1
keywords = ("project timeline template, project management excel, gantt chart template, WBS template, "
            "project budget tracker, risk register excel, task tracker spreadsheet, critical path analysis, "
            "project dashboard, resource allocation template, monthly progress tracker, team velocity, "
            "earned value management, project portfolio, PlanoraNest")
mr(ws_abt, r, "Keywords / SEO Tags", Font(name='Calibri', size=13, bold=True, color=DARK_BLUE), 24)
r += 1
mr(ws_abt, r, keywords, Font(name='Calibri', size=8, color='999999'), 30)
r += 2
mr(ws_abt, r, f"(c) {YR} PlanoraNest. All rights reserved. | {SITE}", Font(name='Calibri', size=9, color='999999'), 20)

print("  [OK] About sheet created")

# ═══════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════
os.makedirs(OUT, exist_ok=True)
wb.save(OUT_FILE)
print("=" * 60)
print(f"[OK] Complex Project Timeline template saved to:")
print(f"     {OUT_FILE}")
print(f"[OK] 7 sheets | 200+ formulas | data validation | conditional formatting | charts")
print("=" * 60)
