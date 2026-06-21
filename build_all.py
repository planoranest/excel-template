#!/usr/bin/env python3
"""
PlanoraNest 240 Professional Excel Template Generator v2.0
Multi-dimensional templates with formula interactions, data validation, conditional formatting, and charts.
https://planoranest.com/ | MIT License
"""
import openpyxl, os, datetime, json, random
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule, ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from copy import copy

OUT = os.path.dirname(os.path.abspath(__file__))
SITE = "https://planoranest.com/"
YR = datetime.datetime.now().year
TD = datetime.date.today().strftime("%B %d, %Y")

# ═══════════════════════════════ STYLES ═══════════════════════════════
NAVY, DARK_BLUE, STEEL, ACCENT = "1B2A4A", "1F4E79", "2E75B6", "4472C4"
LIGHT_BG, DARK_BG, WHITE_BG = "F2F7FB", "E8F0FE", "FFFFFF"
GREEN, RED, ORANGE, YELLOW, PURPLE, GRAY = "1B7A3D", "C0392B", "E67E22", "F1C40F", "7D3C98", "95A5A6"
DARK_TEXT, MED_TEXT = "1A1A1A", "555555"

thin_border = Border(left=Side('thin','D5D8DC'), right=Side('thin','D5D8DC'),
                     top=Side('thin','D5D8DC'), bottom=Side('thin','D5D8DC'))
header_border = Border(left=Side('thin',NAVY), right=Side('thin',NAVY),
                       top=Side('thin',NAVY), bottom=Side('medium',NAVY))
card_border = Border(left=Side('medium',STEEL), right=Side('thin','E5E7E9'),
                     top=Side('thin','E5E7E9'), bottom=Side('thin','E5E7E9'))

hdr_fill  = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
hdr_font  = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
title_font = Font(name='Calibri', size=18, bold=True, color=NAVY)
sub_font  = Font(name='Calibri', size=11, italic=True, color=MED_TEXT)
sec_font  = Font(name='Calibri', size=12, bold=True, color=DARK_BLUE)
data_font = Font(name='Calibri', size=10, color=DARK_TEXT)
bold_font = Font(name='Calibri', size=10, bold=True, color=NAVY)
link_font = Font(name='Calibri', size=10, color='0563C1', underline='single')
small_font = Font(name='Calibri', size=8, color='999999')
note_font  = Font(name='Calibri', size=9, italic=True, color=MED_TEXT)
kpi_val_font = Font(name='Calibri', size=22, bold=True, color=DARK_BLUE)
kpi_lbl_font = Font(name='Calibri', size=9, color=MED_TEXT)
code_font = Font(name='Consolas', size=9, color='777777')
stat_font = Font(name='Calibri', size=14, bold=True, color=DARK_BLUE)

alt_fill   = PatternFill(start_color=LIGHT_BG, end_color=LIGHT_BG, fill_type='solid')
green_fill = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
red_fill   = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
yellow_fill = PatternFill(start_color='FCF3CF', end_color='FCF3CF', fill_type='solid')
blue_fill  = PatternFill(start_color='D6EAF8', end_color='D6EAF8', fill_type='solid')
card_fill  = PatternFill(start_color=WHITE_BG, end_color=WHITE_BG, fill_type='solid')
light_fills = [None, alt_fill]

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left   = Alignment(horizontal='left', vertical='center', wrap_text=True)
right  = Alignment(horizontal='right', vertical='center')

def ac(cell, **kw):
    for k,v in kw.items(): setattr(cell, k, v)

# ═══════════════════════════ CORE HELPERS ═══════════════════════════
def write_cell(ws, r, c, v, font=data_font, fill=None, align=center, border=thin_border, nf=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font; cell.alignment = align; cell.border = border
    if fill: cell.fill = fill
    if nf: cell.number_format = nf
    return cell

def write_hdr_row(ws, r, headers, sc=1):
    for i,h in enumerate(headers):
        write_cell(ws, r, sc+i, h, font=hdr_font, fill=hdr_fill, align=center, border=header_border)
    ws.row_dimensions[r].height = 30

def write_data_row(ws, r, vals, sc=1, fonts=None, fills=None, aligns=None, nfs=None):
    rf = alt_fill if r % 2 == 0 else None
    for i,v in enumerate(vals):
        f  = fonts[i] if fonts and i < len(fonts) and fonts[i] else data_font
        fl = fills[i] if fills and i < len(fills) and fills[i] is not None else rf
        a  = aligns[i] if aligns and i < len(aligns) and aligns[i] else center
        n  = nfs[i] if nfs and i < len(nfs) and nfs[i] else None
        write_cell(ws, r, sc+i, v, font=f, fill=fl, align=a, nf=n)

def set_widths(ws, widths, sc=1):
    for i,w in enumerate(widths):
        ws.column_dimensions[get_column_letter(sc+i)].width = w

def sec_hdr(ws, r, c1, cn, text):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=cn)
    write_cell(ws, r, c1, text, font=sec_font,
               fill=PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid'),
               align=left, border=Border())
    ws.row_dimensions[r].height = 26

def merge_title_row(ws, r, c1, cn, text):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=cn)
    write_cell(ws, r, c1, text, font=title_font, align=left, border=Border())
    ws.row_dimensions[r].height = 36

def mr(ws, r, v, font, h=24):
    ws.merge_cells(f'A{r}:B{r}')
    ws.cell(row=r, column=1, value=v).font = font
    ws.cell(row=r, column=1).alignment = left
    ws.row_dimensions[r].height = h

def add_dv(ws, col_letter, row_start, row_end, options_list, allow_blank=True):
    dv = DataValidation(type="list", formula1=f'"{",".join(options_list)}"', allow_blank=allow_blank)
    ws.add_data_validation(dv)
    for r in range(row_start, row_end+1):
        dv.add(ws.cell(row=r, column=openpyxl.utils.column_index_from_string(col_letter)))
    return dv

# ═══════════════════════════ SHEET BUILDERS ═══════════════════════════
def make_template_sheet(ws, title, subtitle, headers, widths, data_rows,
                         dv_cols=None, cf_rules=None, summary_rows=None, freeze_at=None):
    """Build a professionally formatted template data sheet."""
    nc = len(headers)
    merge_title_row(ws, 1, 1, nc, title)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=nc)
    write_cell(ws, 2, 1, subtitle, font=sub_font, align=left, border=Border())
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 6

    hr = 4
    write_hdr_row(ws, hr, headers)
    set_widths(ws, widths)

    dr_start = hr + 1
    for i, row_vals in enumerate(data_rows):
        write_data_row(ws, dr_start + i, row_vals)

    dr_end = dr_start + len(data_rows) - 1

    # Data validations
    if dv_cols:
        for col_letter, options in dv_cols.items():
            add_dv(ws, col_letter, dr_start, max(dr_end, dr_start+50), options)

    # Conditional formatting
    if cf_rules:
        for rule in cf_rules:
            rng = f'{rule["col_letter"]}{dr_start}:{rule["col_letter"]}{max(dr_end, dr_start+50)}'
            if rule['type'] == 'databar':
                ws.conditional_formatting.add(rng, DataBarRule(start_type='min', end_type='max', color=rule.get('color', STEEL), showValue=True))
            elif rule['type'] == 'colorscale':
                ws.conditional_formatting.add(rng, ColorScaleRule(start_type='min', start_color=rule.get('lo','1B7A3D'),
                    mid_type='percentile', mid_value=50, mid_color=rule.get('mid','F1C40F'),
                    end_type='max', end_color=rule.get('hi','C0392B')))
            elif rule['type'] == 'cell_match':
                for match_val, match_fill in rule.get('matches', []):
                    ws.conditional_formatting.add(rng, CellIsRule(operator='equal', formula=[f'"{match_val}"'],
                        fill=PatternFill(start_color=match_fill, end_color=match_fill, fill_type='solid')))

    # Summary rows
    if summary_rows:
        sr = dr_end + 2
        sec_hdr(ws, sr, 1, nc, summary_rows.get('title', 'Summary'))
        for si, srow in enumerate(summary_rows.get('rows', [])):
            row = sr + 1 + si
            for ci, cv in enumerate(srow):
                write_cell(ws, row, ci+1, cv, font=bold_font if ci == 0 else data_font, align=left if ci == 0 else center)

    if freeze_at:
        ws.freeze_panes = freeze_at
    return dr_start, dr_end

def make_instructions_sheet(ws, template_title, overview, steps_groups, tips, formulas=None):
    ws.sheet_properties.tabColor = STEEL
    ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 95

    mr(ws, 1, f"How to Use: {template_title}", Font(name='Calibri', size=18, bold=True, color=NAVY), 38)
    mr(ws, 3, "Overview", Font(name='Calibri', size=13, bold=True, color=DARK_BLUE), 24)
    mr(ws, 4, overview, data_font, 48)

    r = 6
    for group_title, group_steps in steps_groups:
        mr(ws, r, group_title, Font(name='Calibri', size=12, bold=True, color=STEEL), 26)
        r += 1
        if isinstance(group_steps, list):
            for s in group_steps:
                mr(ws, r, f"• {s}", data_font, 22)
                r += 1
        else:
            mr(ws, r, group_steps, data_font, 22)
            r += 1
        r += 1

    if tips:
        mr(ws, r, "Pro Tips", Font(name='Calibri', size=12, bold=True, color=STEEL), 26)
        r += 1
        for t in tips:
            mr(ws, r, f"💡 {t}", data_font, 24)
            r += 1
        r += 1

    if formulas:
        mr(ws, r, "Key Formulas Used", Font(name='Calibri', size=12, bold=True, color=STEEL), 26)
        r += 1
        for f in formulas:
            mr(ws, r, f"  = {f}", code_font, 19)
            r += 1

def make_about_sheet(ws, template_title, category, keywords, extra_info=None):
    ws.sheet_properties.tabColor = NAVY
    ws.column_dimensions['A'].width = 5; ws.column_dimensions['B'].width = 95

    mr(ws, 1, f"About: {template_title}", Font(name='Calibri', size=18, bold=True, color=NAVY), 38)

    items = [
        ("Template Name", template_title), ("Category", category), ("Version", "2.0 Professional"),
        ("Last Updated", TD), ("Compatibility", "Excel 2010+, Google Sheets, LibreOffice Calc"),
        ("License", "MIT - Free for personal & commercial use"), ("Website", SITE),
    ]
    for i, (lb, vl) in enumerate(items, 3):
        ws.cell(row=i, column=1, value=f"{lb}:").font = bold_font
        f = link_font if lb == "Website" else data_font
        ws.cell(row=i, column=2, value=vl).font = f
        if lb == "Website": ws.cell(row=i, column=2).hyperlink = SITE
        ws.row_dimensions[i].height = 22

    r = 13
    if extra_info:
        mr(ws, r, "What's Included", Font(name='Calibri', size=12, bold=True, color=DARK_BLUE), 24)
        r += 1
        for ei in extra_info:
            mr(ws, r, f"  • {ei}", data_font, 21)
            r += 1
        r += 1

    mr(ws, r, "Keywords / SEO Tags", Font(name='Calibri', size=12, bold=True, color=DARK_BLUE), 24)
    r += 1
    mr(ws, r, keywords, small_font, 30)
    r += 2
    mr(ws, r, f"(c) {YR} PlanoraNest. All rights reserved. | {SITE}", Font(name='Calibri', size=9, color='999999'), 20)

# ═══════════════════════════ DOMAIN DATA GENERATORS ═══════════════════════════
# Each returns: (headers, widths, data_rows, dv_cols, cf_rules, summary_rows, freeze_at)

def R(*vals): return list(vals)

# ─── FINANCE: Portfolio / Investment ───
def gen_portfolio_tracker(title, items):
    hdrs = ["Ticker","Company","Buy Date","Shares","Buy Price($)","Current($)","Invested($)","Value($)","Gain/Loss($)","Return(%)","Alloc(%)","Sector","Strategy","Notes"]
    wids = [8,18,11,8,12,12,12,12,12,10,10,14,14,16]
    data = []
    for i, it in enumerate(items):
        rn = i+5
        inv = f'=D{rn}*E{rn}'; val = f'=D{rn}*F{rn}'; gl = f'=H{rn}-G{rn}'
        ret = f'=IF(G{rn}>0,(F{rn}-E{rn})/E{rn}*100,"")'
        alloc = f'=IF(SUM(H5:H{rn+len(items)+4})=0,"",H{rn}/SUM(H5:H{rn+len(items)+4}))'
        data.append(R(it[0],it[1],it[2],it[3],it[4],it[5],inv,val,gl,ret,alloc,it[6],it[7],it[8]))
    summary = {'title':'📊 PORTFOLIO SUMMARY', 'rows':[
        R("Total Invested", f'=SUM(G5:G{4+len(data)})',"","", f'Total Value', f'=SUM(H5:H{4+len(data)})',"","", f'Total G/L', f'=SUM(I5:I{4+len(data)})'),
        R("Best Performer","","","", "Worst Performer","","","", "Avg Return %", f'=AVERAGE(J5:J{4+len(data)})/100'),
    ]}
    cf = [{'col_letter':'I','type':'colorscale'},{'col_letter':'J','type':'colorscale'}]
    dv = {'H':['Technology','ETF','Financial','Healthcare','Energy','Real Estate','Consumer','Other']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── FINANCE: Budget / Expense ───
def gen_budget_planner(cats, income_items, expense_items):
    hdrs = ["Category","Budget($)","Actual($)","Variance($)","% of Budget","Trend","Notes"]
    wids = [24,14,14,14,12,12,24]
    data = []
    data.append(R("💰 INCOME","","","","","",""))
    r_offset = 5
    for i, inc in enumerate(income_items):
        rn = r_offset + len(data)
        var = f'=IF(B{rn}="","",C{rn}-B{rn})'
        pct = f'=IF(B{rn}=0,"",C{rn}/B{rn}*100)'
        data.append(R(f"  {inc[0]}", inc[1], inc[2], var, pct, inc[3] if len(inc)>3 else "", inc[4] if len(inc)>4 else ""))
    ti_row = r_offset + len(data)
    data.append(R("TOTAL INCOME",f'=SUM(B{r_offset+1}:B{ti_row-1})',f'=SUM(C{r_offset+1}:C{ti_row-1})',
                  f'=C{ti_row}-B{ti_row}',f'=IF(B{ti_row}=0,"",C{ti_row}/B{ti_row}*100)',"",""))
    data.append(R("💸 EXPENSES","","","","","",""))
    exp_start = r_offset + len(data)
    for i, exp in enumerate(expense_items):
        rn = r_offset + len(data)
        var = f'=IF(B{rn}="","",C{rn}-B{rn})'
        pct = f'=IF(B{rn}=0,"",C{rn}/B{rn}*100)'
        data.append(R(f"  {exp[0]}", exp[1], exp[2], var, pct, exp[3] if len(exp)>3 else "", exp[4] if len(exp)>4 else ""))
    te_row = r_offset + len(data)
    data.append(R("TOTAL EXPENSES",f'=SUM(B{exp_start}:B{te_row-1})',f'=SUM(C{exp_start}:C{te_row-1})',
                  f'=C{te_row}-B{te_row}',f'=IF(B{te_row}=0,"",C{te_row}/B{te_row}*100)',"",""))
    net_row = r_offset + len(data)
    data.append(R("📊 NET","","",f'=C{ti_row}-C{te_row}',"","","← Positive = Saving!"))
    summary = {'title':'📊 SPENDING INSIGHTS', 'rows':[
        R("Savings Rate", f'=IF(C{ti_row}=0,"",C{net_row}/C{ti_row})', "", "",
          "Largest Expense", f'=INDEX(A{exp_start}:A{te_row-1},MATCH(MAX(B{exp_start}:B{te_row-1}),B{exp_start}:B{te_row-1},0))'),
    ]}
    cf = [{'col_letter':'D','type':'colorscale'},{'col_letter':'E','type':'databar'}]
    return hdrs, wids, data, {}, cf, summary, 'A5'

# ─── FINANCE: Calculator type ───
def gen_calculator(title, inputs, calculations, result_label="Result"):
    hdrs = ["Parameter","Value","Unit","Notes"]
    wids = [34,22,12,28]
    data = []
    in_offset = 5
    for i, inp in enumerate(inputs):
        data.append(R(inp[0], inp[1], inp[2] if len(inp)>2 else "", inp[3] if len(inp)>3 else ""))
    divider = in_offset + len(data)
    data.append(R("══════════════════","","",""))
    data.append(R("📊 CALCULATIONS","","",""))
    calc_start = in_offset + len(data)
    for i, calc in enumerate(calculations):
        data.append(R(calc[0], calc[1], calc[2] if len(calc)>2 else "", calc[3] if len(calc)>3 else ""))
    result_row = in_offset + len(data)
    data.append(R("","","",""))
    data.append(R(f"🎯 {result_label}:", calculations[-1][1] if calculations else "=0", "", "← Key decision metric"))
    # Add scenario comparison
    data.append(R("","","",""))
    data.append(R("📈 SCENARIO ANALYSIS","","",""))
    scn_start = in_offset + len(data)
    data.append(R("Optimistic Scenario","","","Change input values above to see impact"))
    data.append(R("Pessimistic Scenario","","","Change input values above to see impact"))
    data.append(R("Most Likely Scenario","","","This is your base case above"))

    dv = {'C':['years','months','$','$/yr','$/mo','%','people','units']}
    cf = [{'col_letter':'B','type':'colorscale','lo':RED,'mid':YELLOW,'hi':GREEN}]
    summary = {'title':'📊 QUICK SUMMARY', 'rows':[
        R("Total Inputs", f'=COUNTA(A5:A{in_offset+len(inputs)-1})',"","",
          "Total Calculations", f'=COUNTA(A{calc_start}:A{result_row-1})'),
    ]}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── TRACKER: General purpose log/tracker ───
def gen_tracker_log(title, items, extra_cols=None):
    if extra_cols is None:
        extra_cols = [("Category",14),("Status",12),("Days Elapsed",12),("Priority",10)]
    base_hdrs = ["Date","Item","Description"]
    base_wids = [12,24,28]
    for ec, ew in extra_cols:
        base_hdrs.append(ec); base_wids.append(ew)
    base_hdrs += ["Completion(%)","Follow-up Date","Notes"]
    base_wids += [12,12,20]

    data = []
    for i, it in enumerate(items):
        rn = i + 5
        days_elapsed = f'=IF(A{rn}="","",NETWORKDAYS(A{rn},TODAY()))'
        completion = f'=IF(OR(K{rn}="",A{rn}=""),"",IF(K{rn}>=TODAY(),MIN(100,MAX(0,(TODAY()-A{rn})/(K{rn}-A{rn})*100)),100))'
        vals = list(it[0:3])
        for j in range(3, 3+len(extra_cols)):
            if j == 5 and len(extra_cols) > 2:  # Days Elapsed column
                vals.append(days_elapsed)
            else:
                vals.append(it[j] if j < len(it) else "")
        vals.append(completion)
        vals.append(it[len(it)-2] if len(it) > len(extra_cols)+3 else (f"2025-0{6+i//3}-{(i%3)*10+15:02d}" if i < 3 else ""))
        vals.append(it[-1] if len(it) > len(extra_cols)+4 else "")
        data.append(R(*vals))

    summary = {'title':'📊 SUMMARY', 'rows':[
        R("Total Entries", f'=COUNTA(B5:B{4+len(data)})',"","",
          "This Month", f'=COUNTIFS(A5:A{4+len(data)},">="&DATE(YEAR(TODAY()),MONTH(TODAY()),1))',"","",
          "Completed", f'=COUNTIF(H5:H{4+len(data)},"Complete")'),
    ]}
    cf = [{'col_letter':'H','type':'cell_match','matches':[
        ('Complete','D5F5E3'),('In Progress','D6EAF8'),('Pending','FCF3CF'),('Blocked','FADBD8')]}]
    dv = {'D':['Category A','Category B','Category C','Category D'],
          'H':['Not Started','In Progress','Complete','Pending','Blocked'],
          'I':['Critical','High','Medium','Low']}
    return base_hdrs, base_wids, data, dv, cf, summary, 'A5'

# ─── PLANNER: Goals / Projects / Planning ───
def gen_planner(title, goals, track_progress=True):
    hdrs = ["Goal / Task","Category","Target","Current","Progress(%)","Start","Deadline","Priority","Status","Owner","Notes"]
    wids = [22,14,12,12,10,11,11,10,12,14,20]
    data = []
    for i, g in enumerate(goals):
        rn = i + 5
        prog = f'=IF(C{rn}=0,"",D{rn}/C{rn})' if track_progress else ""
        data.append(R(g[0],g[1],g[2] if g[2] else "",g[3] if g[3] else "",
                      prog,g[4] if len(g)>4 else "",g[5] if len(g)>5 else "",
                      g[6] if len(g)>6 else "Medium",g[7] if len(g)>7 else "In Progress",
                      g[8] if len(g)>8 else "",g[9] if len(g)>9 else ""))
    summary = {'title':'📊 PROGRESS OVERVIEW', 'rows':[
        R("Total Goals", f'=COUNTA(A5:A{4+len(data)})',"","",
          "Completed", f'=COUNTIF(I5:I{4+len(data)},"Complete")',"","",
          "Avg Progress", f'=AVERAGE(E5:E{4+len(data)})'),
    ]}
    cf = [{'col_letter':'E','type':'databar'},
          {'col_letter':'I','type':'cell_match','matches':[
              ('Complete','D5F5E3'),('In Progress','D6EAF8'),('Not Started','EAECEE'),
              ('Blocked','FADBD8'),('On Hold','FDEBD0')]}]
    dv = {'H':['Critical','High','Medium','Low'],'I':['Not Started','In Progress','Complete','Blocked','On Hold']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── ANALYSIS: SWOT / Comparison / Assessment ───
def gen_analysis_matrix(title, categories, criteria, items_data):
    # criteria as columns, items as rows, with scoring
    hdrs = ["Item"] + [c[0] for c in criteria] + ["Total Score","Rank","Comments"]
    wids = [22] + [12]*len(criteria) + [12,8,18]
    data = []
    for i, it in enumerate(items_data):
        rn = i + 5
        col_start = 2  # B
        col_end = 1 + len(criteria)  # last criteria col
        total_col_letter = get_column_letter(1 + len(criteria) + 1)  # Total Score column
        sum_parts = "+".join([f"{get_column_letter(col_start+j)}{rn}" for j in range(len(criteria))])
        score = f'={sum_parts}'
        rank = f'=IF({total_col_letter}{rn}="","",RANK({total_col_letter}{rn},${total_col_letter}$5:${total_col_letter}${4+len(items_data)}))'
        data.append(R(it[0], *it[1:1+len(criteria)], score, rank, it[-1] if len(it) > 1+len(criteria) else ""))
    cf = [{'col_letter': get_column_letter(1+len(criteria)+1), 'type':'colorscale'}]
    return hdrs, wids, data, {}, cf, {'title':'📊 RANKING'}, 'A5'

# ─── IT / Technical ───
def gen_it_asset(title, assets):
    hdrs = ["Asset ID","Asset Name","Type","Department","Purchase Date","Cost($)","Depreciation(%)","Current Value($)","Warranty Exp","Assigned To","Status","Location","Notes"]
    wids = [10,20,12,14,12,12,12,14,12,16,12,14,16]
    data = []
    for i, a in enumerate(assets):
        rn = i + 5
        cv = f'=IF(G{rn}="","",F{rn}*(1-G{rn}/100))'
        data.append(R(a[0],a[1],a[2],a[3],a[4],a[5],a[6],cv,a[7] if len(a)>7 else "",
                      a[8] if len(a)>8 else "",a[9] if len(a)>9 else "Active",
                      a[10] if len(a)>10 else "",a[11] if len(a)>11 else ""))
    summary = {'title':'📊 ASSET SUMMARY', 'rows':[
        R("Total Assets", f'=COUNTA(A5:A{4+len(data)})',"","",
          "Total Value", f'=SUM(H5:H{4+len(data)})',"","",
          "Active", f'=COUNTIF(K5:K{4+len(data)},"Active")'),
    ]}
    cf = [{'col_letter':'H','type':'databar'},
          {'col_letter':'K','type':'cell_match','matches':[('Active','D5F5E3'),('Retired','FADBD8'),('Maintenance','FCF3CF')]}]
    dv = {'C':['Hardware','Software','Network','Peripheral','Cloud','License'],'K':['Active','Retired','Maintenance','Lost']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── Product Management ───
def gen_product_template(title, items, extra_type="roadmap"):
    if extra_type == "roadmap":
        hdrs = ["Feature","Theme","RICE Score","Reach","Impact(1-5)","Confidence(%)","Effort(Wks)","Quarter","Status","Dependencies","Owner","Notes"]
        wids = [22,14,10,8,10,12,10,10,12,18,14,18]
    elif extra_type == "launch":
        hdrs = ["Task","Category","Owner","Start Date","Due Date","Duration(Days)","Dependencies","Status","% Complete","Blockers","Checklist Item","Notes"]
        wids = [20,14,14,11,11,11,14,12,10,18,20,18]
    elif extra_type == "metrics":
        hdrs = ["Metric","Category","Current","Target","Previous","Change(%)","Trend","Frequency","Owner","Source","Benchmark","Notes"]
        wids = [18,14,12,12,12,10,10,12,14,14,12,16]
    else:
        hdrs = ["Item","Category","Priority","Status","Owner","Start","End","Progress(%)","Value","Impact","Notes"]
        wids = [22,14,10,12,14,10,10,10,14,14,18]

    data = []
    for i, it in enumerate(items):
        rn = i + 5
        vals = list(it)
        # Add formulas for RICE score, duration, or progress as appropriate
        if extra_type == "roadmap" and len(vals) >= 7:
            # Auto-calc RICE score = Reach * Impact * Confidence / Effort
            rice_formula = f'=IF(OR(D{rn}="",E{rn}="",F{rn}="",G{rn}="",G{rn}=0),"",ROUND(D{rn}*E{rn}*F{rn}/G{rn},1))'
            row_vals = vals[0:2] + [rice_formula] + [vals[j] if j < len(vals) else "" for j in range(3, min(12, len(vals)+1))]
            row_vals = (row_vals + [""]*12)[:12]
            data.append(R(*row_vals))
        elif extra_type == "launch" and len(vals) >= 6:
            dur_formula = f'=IF(AND(D{rn}<>"",E{rn}<>""),NETWORKDAYS(D{rn},E{rn}),"")'
            row_vals = vals[0:5] + [dur_formula] + vals[5:11]
            row_vals = (row_vals + [""]*12)[:12]
            data.append(R(*row_vals))
        elif extra_type == "metrics" and len(vals) >= 6:
            change_formula = f'=IF(OR(C{rn}="",E{rn}="",E{rn}=0),"",ROUND((C{rn}-E{rn})/E{rn}*100,1))'
            trend_formula = f'=IF(F{rn}="","",IF(F{rn}>0,"📈 Up",IF(F{rn}<0,"📉 Down","➡️ Flat")))'
            row_vals = vals[0:5] + [change_formula, trend_formula] + vals[6:11]
            row_vals = (row_vals + [""]*12)[:12]
            data.append(R(*row_vals))
        else:
            padded = (vals + [""]*12)[:12]
            data.append(R(*padded))

    summary = {'title':'📊 OVERVIEW', 'rows':[
        R("Total Items", f'=COUNTA(A5:A{4+len(data)})',"","",
          "In Progress", f'=COUNTIF(D5:D{4+len(data)},"In Progress")',"","",
          "Completed", f'=COUNTIF(D5:D{4+len(data)},"Complete")'),
    ]}
    cf = [{'col_letter':'C' if extra_type=='roadmap' else 'H','type':'databar'},
          {'col_letter':'I' if extra_type=='roadmap' else 'D','type':'cell_match','matches':[
              ('Complete','D5F5E3'),('In Progress','D6EAF8'),('Planned','EAECEE'),('Blocked','FADBD8'),('At Risk','FDEBD0')]}]
    dv = {'D':['Not Started','In Progress','Complete','Blocked','At Risk','Planned'],
          'C':['Critical','High','Medium','Low']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── Health / Fitness ───
def gen_health_tracker(title, entries):
    hdrs = ["Date","Activity/Meal","Duration(mins)","Calories","Distance(km)","Heart Rate","Weight(kg)","Mood","Category","Intensity","Cal/Min","Notes"]
    wids = [11,22,11,10,10,10,10,8,13,9,9,18]
    data = []
    for i, e in enumerate(entries):
        rn = i + 5
        cal_per_min = f'=IF(AND(C{rn}<>"",D{rn}<>"",C{rn}>0),ROUND(D{rn}/C{rn},1),"")'
        vals = list(e)
        row_vals = vals[0:9] + [vals[9] if len(vals)>9 else "Moderate", cal_per_min, vals[10] if len(vals)>10 else ""]
        if len(row_vals) < 12:
            row_vals += [""] * (12 - len(row_vals))
        data.append(R(*row_vals[0:12]))

    summary = {'title':'📊 WEEKLY ANALYTICS', 'rows':[
        R("Total Duration", f'=SUM(C5:C{4+len(data)})',"mins","",
          "Total Calories", f'=SUM(D5:D{4+len(data)})',"kcal","",
          "Avg Heart Rate", f'=AVERAGE(F5:F{4+len(data)})',"bpm","",
          "Cal/Min Avg", f'=AVERAGE(K5:K{4+len(data)})'),
    ]}
    cf = [{'col_letter':'D','type':'databar','color':GREEN},
          {'col_letter':'H','type':'colorscale','lo':RED,'mid':YELLOW,'hi':GREEN},
          {'col_letter':'K','type':'colorscale','lo':RED,'hi':GREEN}]
    dv = {'I':['Cardio','Strength','Yoga','Running','Cycling','Swimming','Walking','HIIT','Sports','Other'],
          'J':['Low','Moderate','High','Intense']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── Education / Learning ───
def gen_edu_tracker(title, items):
    hdrs = ["Course/Topic","Provider","Category","Start Date","End Date","Duration(Days)","Progress(%)","Grade","Status","Credits/Cert","Cost($)","ROI Rating","Notes"]
    wids = [22,15,12,11,11,11,10,8,12,11,9,10,18]
    data = []
    for i, it in enumerate(items):
        rn = i + 5
        dur = f'=IF(AND(D{rn}<>"",E{rn}<>""),NETWORKDAYS(D{rn},E{rn}),"")'
        roi = f'=IF(AND(G{rn}<>"",G{rn}>50,K{rn}<>""),IF(K{rn}>0,ROUND(G{rn}*100/K{rn},1),""),"")'
        vals = list(it)
        row_vals = vals[0:5] + [dur, vals[5] if len(vals)>5 else "", vals[6] if len(vals)>6 else "",
                  vals[7] if len(vals)>7 else "In Progress", vals[8] if len(vals)>8 else "",
                  vals[9] if len(vals)>9 else "", roi, vals[10] if len(vals)>10 else ""]
        if len(row_vals) < 13:
            row_vals += [""] * (13 - len(row_vals))
        data.append(R(*row_vals[0:13]))

    summary = {'title':'📊 LEARNING ANALYTICS', 'rows':[
        R("Total Courses", f'=COUNTA(A5:A{4+len(data)})',"","",
          "Completed", f'=COUNTIF(I5:I{4+len(data)},"Complete")',"","",
          "Total Spent", f'=SUM(K5:K{4+len(data)})',"","",
          "Avg Progress", f'=AVERAGE(G5:G{4+len(data)})/100'),
    ]}
    cf = [{'col_letter':'G','type':'databar'},
          {'col_letter':'L','type':'colorscale','lo':RED,'hi':GREEN},
          {'col_letter':'I','type':'cell_match','matches':[('Complete','D5F5E3'),('In Progress','D6EAF8'),('Planned','EAECEE'),('Dropped','FADBD8')]}]
    dv = {'C':['Technical','Soft Skills','Business','Creative','Language','Certification','Other'],
          'I':['Planned','In Progress','Complete','Dropped','On Hold']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── Life / Lifestyle ───
def gen_life_tracker(title, entries, mood_tracking=False):
    if mood_tracking:
        hdrs = ["Date","Mood(1-10)","Energy(1-10)","Activity","People","Gratitude","Challenge","Lesson Learned","Self-Care?","Sleep(Hrs)","Notes"]
        wids = [10,10,10,20,16,26,22,22,12,10,18]
    else:
        hdrs = ["Date","Event","Category","People","Location","Rating","Cost($)","Duration(Hrs)","Value Score","Follow-up?","Notes"]
        wids = [10,22,14,14,14,8,8,10,10,10,18]
    data = []
    for i, e in enumerate(entries):
        rn = i + 5
        vals = list(e)
        if mood_tracking:
            padded = list(vals[0:11]) if len(vals) >= 11 else list(vals) + [""] * (11 - len(vals))
            data.append(R(*padded))
        else:
            value_score = f'=IF(AND(F{rn}<>"",G{rn}<>"",G{rn}>0),ROUND(F{rn}/G{rn}*10,1),"")'
            row_vals = vals[0:7] if len(vals) >= 7 else vals + [""]*(7-len(vals))
            row_vals = row_vals[:7] + [vals[7] if len(vals)>7 else "", value_score] + [vals[8] if len(vals)>8 else "", vals[9] if len(vals)>9 else ""]
            padded = (row_vals + [""]*11)[:11]
            data.append(R(*padded))

    summary = {'title':'📊 WEEKLY INSIGHTS', 'rows':[
        R("Total Events", f'=COUNTA(B5:B{4+len(data)})',"","",
          f"Avg {'Mood' if mood_tracking else 'Rating'}", f'=AVERAGE({("B" if mood_tracking else "F")}5:{("B" if mood_tracking else "F")}{4+len(data)})',"","",
          "Total Cost" if not mood_tracking else "Best Day",
          f'={("SUM(G5:G" if not mood_tracking else "INDEX(A5:A")}{4+len(data)}){")" if not mood_tracking else ",MATCH(MAX(B5:B"+str(4+len(data))+"),B5:B"+str(4+len(data))+",0))"}'),
    ]}
    cf = [{'col_letter':'B' if mood_tracking else 'F','type':'colorscale','lo':RED,'mid':YELLOW,'hi':GREEN}] + \
         ([{'col_letter':'I','type':'colorscale','lo':RED,'hi':GREEN}] if not mood_tracking else [])
    dv = {'C':['Exercise','Social','Work','Family','Hobby','Self-Care','Chores','Entertainment','Learning','Travel']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── Home / Family ───
def gen_home_tracker(title, items):
    hdrs = ["Item/Room","Category","Frequency","Last Done","Next Due","Days Until","Assigned To","Status","Priority","Cost($)","Supplies Needed","Notes"]
    wids = [18,14,12,11,11,10,14,12,10,10,20,18]
    data = []
    for i, it in enumerate(items):
        rn = i + 5
        days_until = f'=IF(E{rn}="","",MAX(0,NETWORKDAYS(TODAY(),E{rn})))'
        data.append(R(it[0],it[1],it[2] if len(it)>2 else "Weekly",
                      it[3] if len(it)>3 else "",it[4] if len(it)>4 else "",
                      days_until,it[5] if len(it)>5 else "",
                      it[6] if len(it)>6 else "Pending",it[7] if len(it)>7 else "Medium",
                      it[8] if len(it)>8 else "",it[9] if len(it)>9 else "",it[10] if len(it)>10 else ""))
    cf = [{'col_letter':'F','type':'colorscale','lo':RED,'mid':YELLOW,'hi':GREEN},
          {'col_letter':'H','type':'cell_match','matches':[('Done','D5F5E3'),('Pending','FCF3CF'),('Overdue','FADBD8')]}]
    dv = {'H':['Done','Pending','Overdue','Skipped'],'G':['Critical','High','Medium','Low']}
    return hdrs, wids, data, dv, cf, {'title':'📊 HOME SUMMARY'}, 'A5'

# ─── Travel ───
def gen_travel_tracker(title, items):
    hdrs = ["Destination/Item","Category","Start Date","End Date","Duration(Days)","Budget($)","Actual($)","Variance($)","Booking Status","Booking Ref","Packing List","Notes/Rating"]
    wids = [20,12,11,11,11,12,12,12,12,14,20,18]
    data = []
    for i, it in enumerate(items):
        rn = i + 5
        dur = f'=IF(AND(C{rn}<>"",D{rn}<>""),NETWORKDAYS(C{rn},D{rn}),"")'
        var = f'=IF(AND(F{rn}<>"",G{rn}<>""),F{rn}-G{rn},"")'
        data.append(R(it[0],it[1],it[2],it[3],dur,it[5] if len(it)>5 else "",
                      it[6] if len(it)>6 else "",var,it[7] if len(it)>7 else "Not Booked",
                      it[8] if len(it)>8 else "",it[9] if len(it)>9 else "",it[10] if len(it)>10 else ""))
    cf = [{'col_letter':'H','type':'colorscale'}]
    dv = {'I':['Booked','Not Booked','Confirmed','Cancelled','Waitlisted']}
    return hdrs, wids, data, dv, cf, {'title':'📊 TRIP SUMMARY'}, 'A5'

# ─── Creative / Hobbies ───
def gen_creative_tracker(title, items):
    hdrs = ["Project/Item","Type","Status","Started","Completed","Duration(Days)","Difficulty","Satisfaction","Cost($)","Materials/Tools","Inspiration","Notes"]
    wids = [20,12,12,11,11,10,9,10,9,18,18,16]
    data = []
    for i, it in enumerate(items):
        rn = i + 5
        dur = f'=IF(AND(D{rn}<>"",E{rn}<>""),NETWORKDAYS(D{rn},E{rn}),IF(D{rn}<>"",NETWORKDAYS(D{rn},TODAY()),""))'
        vals = list(it)
        if len(vals) > 4:
            # Insert duration formula between completed and difficulty
            row_vals = vals[0:5] + [dur] + vals[5:10] if len(vals) > 5 else vals[0:5] + [dur]
            row_vals = row_vals[0:12]
            if len(row_vals) < 12:
                row_vals += [""] * (12 - len(row_vals))
            data.append(R(*row_vals))
        else:
            data.append(R(*vals[0:5], dur, *vals[5:]))

    summary = {'title':'📊 CREATIVE STATS', 'rows':[
        R("Total Projects", f'=COUNTA(A5:A{4+len(data)})',"","",
          "Completed", f'=COUNTIF(C5:C{4+len(data)},"Complete")',"","",
          "Total Cost", f'=SUM(I5:I{4+len(data)})',"","",
          "Avg Satisfaction", f'=AVERAGE(H5:H{4+len(data)})'),
    ]}
    cf = [{'col_letter':'G','type':'colorscale','lo':GREEN,'hi':RED},
          {'col_letter':'H','type':'databar','color':PURPLE},
          {'col_letter':'C','type':'cell_match','matches':[('Complete','D5F5E3'),('In Progress','D6EAF8'),('Idea','FCF3CF'),('On Hold','FADBD8')]}]
    dv = {'B':['Painting','Writing','Music','Photography','DIY','Craft','Woodworking','Knitting','Digital Art','Gardening','Other'],
          'C':['Idea','Planning','In Progress','Complete','On Hold','Abandoned']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ─── Productivity ───
def gen_productivity_tracker(title, items):
    hdrs = ["Task/Activity","Category","Priority","Est(Hrs)","Actual(Hrs)","Efficiency(%)","Focus(1-10)","Distractions","Done?","Date","Method","Reflection"]
    wids = [20,13,9,9,9,10,9,14,8,10,13,18]
    data = []
    for i, it in enumerate(items):
        rn = i + 5
        eff = f'=IF(AND(D{rn}<>"",E{rn}<>"",E{rn}>0),MIN(200,D{rn}/E{rn}*100),"")'
        vals = list(it)
        # Insert efficiency formula
        if len(vals) >= 5:
            row_vals = vals[0:5] + [eff] + vals[5:11]
        else:
            row_vals = vals[0:5] + [eff]
        if len(row_vals) < 12:
            row_vals += [""] * (12 - len(row_vals))
        data.append(R(*row_vals[0:12]))

    summary = {'title':'📊 PRODUCTIVITY ANALYTICS', 'rows':[
        R("Total Tasks", f'=COUNTA(A5:A{4+len(data)})',"","",
          "Completed", f'=COUNTIF(I5:I{4+len(data)},"Yes")',"","",
          "Avg Efficiency", f'=AVERAGE(F5:F{4+len(data)})/100',"","",
          "Avg Focus", f'=AVERAGE(G5:G{4+len(data)})',"/10"),
    ]}
    cf = [{'col_letter':'F','type':'databar','color':STEEL},
          {'col_letter':'G','type':'colorscale','lo':RED,'hi':GREEN},
          {'col_letter':'C','type':'cell_match','matches':[('Critical','FADBD8'),('High','FDEBD0'),('Medium','FCF3CF'),('Low','EAECEE')]}]
    dv = {'C':['Critical','High','Medium','Low'],'K':['Pomodoro','GTD','Time Blocking','Deep Work','Eat the Frog','Batch Processing','Kanban']}
    return hdrs, wids, data, dv, cf, summary, 'A5'

# ═══════════════════════════ MASTER TEMPLATE DATABASE ═══════════════════════════
CATS = {
    "01_Investment_Finance": "Investment & Finance",
    "02_Personal_Planning": "Personal Planning & Goals",
    "03_Business_Management": "Business & Management",
    "04_Health_Fitness": "Health & Fitness",
    "05_Education_Learning": "Education & Learning",
    "06_Life_Lifestyle": "Life & Lifestyle",
    "07_Home_Family": "Home & Family",
    "08_Travel_Adventure": "Travel & Adventure",
    "09_Creative_Hobbies": "Creative & Hobbies",
    "10_Productivity_Time": "Productivity & Time Management",
    "11_IT_Technology": "IT & Technology",
    "12_Product_Management": "Product Management",
}

ALL_TEMPLATES = []  # (cat_dir, filename, title, subtitle, gen_func_key, gen_data, overview, steps, tips, formulas, keywords)

def add_template(cat, fn, ttl, sub, gen_key, gen_data, overview, steps, tips, formulas, kw):
    ALL_TEMPLATES.append((cat, fn, ttl, sub, gen_key, gen_data, overview, steps, tips, formulas, kw))

def financial_keywords(title):
    return f"{title.lower()},investment,finance,excel template,spreadsheet,portfolio,budget,calculator,financial planning,professional,free,PlanoraNest"

# ═══════════════════════════ CATEGORY 01: INVESTMENT & FINANCE (20) ═══════════════════════════
C1 = "01_Investment_Finance"

add_template(C1, "01_Stock_Portfolio_Tracker", "Stock Portfolio Tracker",
    "Advanced investment portfolio with real-time gain/loss, return %, sector allocation & diversification analysis",
    "portfolio", [
        ("AAPL","Apple Inc.","2024-01-15",50,185,220.50,"Technology","Growth","Long-term core holding"),
        ("MSFT","Microsoft Corp.","2024-03-10",30,415,448.50,"Technology","Growth","AI & cloud growth play"),
        ("VTI","Vanguard Total Stock","2024-06-01",20,260,278.50,"ETF","Core","Broad market index"),
        ("GOOGL","Alphabet Inc.","2024-04-20",15,175,195.50,"Technology","Growth","Cloud, AI, advertising"),
        ("AMZN","Amazon.com","2024-05-15",25,180,195.30,"Consumer","Growth","E-commerce & AWS"),
        ("BRK.B","Berkshire Hathaway","2024-02-01",10,420,455.80,"Financial","Value","Diversified holding co"),
        ("JNJ","Johnson & Johnson","2024-07-01",20,155,162.40,"Healthcare","Value","Dividend aristocrat"),
        ("NVDA","NVIDIA Corp.","2024-08-15",40,120,155.20,"Technology","Growth","AI chip leader"),
        ("TSLA","Tesla Inc.","2024-06-20",35,245,255.60,"Consumer","Speculative","EV & energy"),
        ("BND","Vanguard Total Bond","2024-03-01",15,72,73.15,"ETF","Income","Fixed income allocation"),
    ],
    "Comprehensive stock portfolio tracker with automatic gain/loss calculations, return percentages, and sector allocation analysis. Track multiple investment positions with real-time market valuations. Ideal for individual investors managing diversified equity portfolios across different sectors and strategies.",
    [("Getting Started",["Enter ticker symbols and company names in the designated columns",
        "Fill in purchase dates, number of shares, and buy prices for cost basis tracking",
        "Update the Current Price column regularly for accurate portfolio valuation",
        "The Invested, Value, Gain/Loss, Return%, and Allocation% columns auto-calculate"]),
     ("Analysis & Strategy",["Review Gain/Loss to identify winning and losing positions",
        "Check Sector column to ensure proper diversification across industries",
        "Monitor Allocation% to avoid over-concentration in any single position",
        "Use the Portfolio Summary section for total portfolio health at a glance"]),
    ],
    ["Update prices at least weekly for accurate valuations",
     "Use conditional formatting colors (green/red) to quickly spot outperformers",
     "Add dividend tracking in the Notes column for total return calculation",
     "Export year-end data to CSV for simplified tax reporting"],
    ["Invested = Shares × Buy_Price","Value = Shares × Current_Price","Gain/Loss = Value − Invested",
     "Return% = (Current − Buy) / Buy × 100","Allocation% = Position / Total Portfolio"],
    financial_keywords("Stock Portfolio Tracker"))

add_template(C1, "02_Monthly_Budget_Planner", "Monthly Budget Planner",
    "Complete household budget with income/expense tracking, budget vs actual analysis & savings rate calculator",
    "budget",
    {"income": [("Salary/Wages",5000,5000,"Stable","Primary income"),
                ("Side Business",800,750,"Variable","Freelance income"),
                ("Investment Income",300,320,"Variable","Dividends & interest"),
                ("Other Income",150,200,"Variable","Miscellaneous")],
     "expenses": [("Housing/Rent",1500,1500,"Fixed",""),
                  ("Utilities (Electric,Water,Gas)",280,265,"Variable",""),
                  ("Groceries & Household",500,480,"Variable",""),
                  ("Transportation (Gas,Parking)",200,185,"Variable",""),
                  ("Insurance (Health,Car,Life)",350,350,"Fixed",""),
                  ("Internet & Phone",120,120,"Fixed",""),
                  ("Entertainment & Dining",300,275,"Variable",""),
                  ("Shopping & Personal Care",200,230,"Variable","Watch this"),
                  ("Savings & Investments",500,500,"Goal","Emergency fund"),
                  ("Debt Repayment",300,300,"Fixed","Student loan")]},
    "Comprehensive monthly budget planner to manage all household income and expenses. Compare budgeted vs actual spending with automatic variance and percentage calculations. Identify overspending patterns, track savings rate, and build financial discipline with clear category breakdowns.",
    [("Setup",["List all income sources with expected monthly amounts",
        "Enter budget allocations for every expense category based on past spending",
        "Record actual spending as transactions occur throughout the month"]),
     ("Review & Adjust",["Review the Variance column to quickly spot over-budget categories",
        "Check % of Budget to identify categories consistently exceeding allocation",
        "Monitor the Net row — green means you're saving, red means deficit",
        "Adjust budget allocations quarterly based on actual spending patterns"])],
    ["Follow the 50/30/20 rule: 50% needs, 30% wants, 20% savings/debt",
     "Track daily expenses via mobile app and update this sheet weekly",
     "Build an emergency fund of 3-6 months of essential expenses first",
     "Use the Savings Rate metric to measure progress toward financial independence"],
    ["Variance = Actual − Budget","% of Budget = Actual / Budget × 100","Savings Rate = Net / Total Income",
     "SUM for category totals","AVERAGE for trend analysis"],
    financial_keywords("Monthly Budget Planner"))

add_template(C1, "03_Retirement_Calculator", "Retirement Savings Calculator",
    "Comprehensive retirement planner with compound interest, inflation adjustment, 4% safe withdrawal rate & Monte Carlo readiness",
    "calculator",
    {"inputs": [("Current Age",30,"years","Your current age"),
                ("Retirement Age",65,"years","Target retirement age"),
                ("Current Savings",50000,"$","Current retirement account balance"),
                ("Monthly Contribution",1500,"$/mo","How much you save each month"),
                ("Expected Annual Return",7,"%","Conservative: 6-7%, Aggressive: 8-10%"),
                ("Expected Inflation Rate",2.5,"%","Historical average ~2-3%"),
                ("Desired Retirement Income",60000,"$/yr","In today's dollars"),
                ("Life Expectancy",90,"years","Plan for longevity")],
     "calculations": [("Years to Retirement","=B2-B1","years",""),
                      ("Future Value at Retirement",f"=FV(B5/12,(B2-B1)*12,-B4,-B3,0)","$","Total nest egg"),
                      ("Inflation-Adjusted Need",f"=B7*(1+B6/100)^(B2-B1)","$/yr","Future dollars needed"),
                      ("Safe Withdrawal (4% Rule)",f"=FV(B5/12,(B2-B1)*12,-B4,-B3,0)*0.04","$/yr","Annual income"),
                      ("Monthly Retirement Income","=B9/12","$/mo",""),
                      ("Shortfall/Surplus","=B10-B8/12","$/mo","Negative = shortfall")],
     "result_label":"Monthly Surplus/Shortfall"},
    "Advanced retirement planning calculator that projects your nest egg using compound interest, adjusts for inflation, and applies the 4% safe withdrawal rule. Run what-if scenarios by adjusting contribution amounts, retirement age, or expected returns to find your optimal savings strategy.",
    [("Enter Your Data",["Fill in your current age and target retirement age",
        "Enter current retirement savings balance across all accounts",
        "Set your monthly contribution amount (include employer match)",
        "Adjust expected return rate based on your portfolio risk tolerance"]),
     ("Analyze Results",["Review the Future Value projection — this is your estimated nest egg",
        "Check the Inflation-Adjusted Need to understand your real retirement expenses",
        "The Shortfall/Surplus row tells you if you're on track or need to adjust",
        "Run scenarios: increase contributions, delay retirement, or adjust returns"])],
    ["Start early — compound interest rewards time more than higher contributions",
     "Include employer 401(k) match in your monthly contribution figure",
     "Revisit this calculator annually and adjust for life changes",
     "Consider healthcare costs which typically rise faster than general inflation"],
    ["FV(rate, nper, pmt, pv, type) — Future value with compound interest",
     "Safe Withdrawal = Nest Egg × 4%","Inflation-Adjusted = Amount × (1+inflation)^years"],
    financial_keywords("Retirement Savings Calculator"))

add_template(C1, "04_Debt_Repayment_Plan", "Debt Repayment Snowball/Avalanche Planner",
    "Strategic debt payoff planner comparing snowball vs avalanche methods with payoff date projections",
    "planner",
    [("Credit Card A","Debt",8000,6500,"2025-01-01","2027-06-30","Critical","In Progress","Self","18.99% APR"),
     ("Credit Card B","Debt",4500,4200,"2025-01-01","2026-12-31","High","In Progress","Self","22.99% APR"),
     ("Student Loan","Debt",35000,32000,"2022-06-01","2030-12-31","High","In Progress","Self","5.5% fixed"),
     ("Car Loan","Debt",18000,15000,"2023-03-15","2027-03-15","Medium","In Progress","Joint","4.9% fixed"),
     ("Personal Loan","Debt",5000,4500,"2024-09-01","2026-09-01","Medium","In Progress","Self","8.5% fixed"),
     ("Mortgage","Debt",250000,245000,"2020-05-01","2050-05-01","Low","In Progress","Joint","3.75% fixed"),
     ("Emergency Fund","Savings",15000,8000,"2024-01-01","2025-12-31","Critical","In Progress","Self","Building 6mo"),
     ("Investment Account","Invest","50000,30000","2023-01-01","2035-01-01","Medium","In Progress","Self","Index funds")],
    "Strategic debt repayment planner that helps you compare the Snowball (smallest balance first) vs Avalanche (highest interest first) methods. Track all debts with current balances, interest rates, and projected payoff dates. Make informed decisions about which debt to prioritize for maximum savings.",
    [("Debt Inventory",["List all debts with current balances and interest rates",
        "Categorize each debt by priority and repayment method",
        "Enter minimum payments and any extra payment amounts"]),
     ("Strategy Execution",["Snowball method: Pay minimum on all debts, put extra toward smallest balance",
        "Avalanche method: Pay minimum on all debts, put extra toward highest interest rate",
        "Update balances monthly to track payoff progress",
        "Celebrate each paid-off debt — momentum is powerful"])],
    ["The avalanche method saves more money mathematically; snowball builds motivation",
     "Consider 0% balance transfer offers but calculate transfer fees carefully",
     "Build a small emergency fund ($1000) before aggressive debt repayment",
     "Automate payments to avoid late fees and protect your credit score"],
    ["Payoff months = NPER(rate/12, -payment, balance, 0)","Total interest = SUM of all interest payments",
     "Snowball order = Sort by balance ascending","Avalanche order = Sort by interest rate descending"],
    financial_keywords("Debt Repayment Planner"))

add_template(C1, "05_Net_Worth_Statement", "Personal Net Worth Statement",
    "Complete net worth tracker with assets, liabilities, automatic net worth calculation & historical trend analysis",
    "portfolio",
    [("Checking Account","Cash","Assets",1,8500,8500,"Liquid","","Primary bank"),
     ("Savings Account","Cash","Assets",1,32000,32000,"Liquid","","Emergency fund"),
     ("401(k) - Employer","Retirement","Assets",1,125000,125000,"Tax-Deferred","","Employer match 6%"),
     ("Roth IRA","Retirement","Assets",1,45000,45000,"Tax-Free","","Vanguard index"),
     ("Brokerage Account","Investment","Assets",1,55000,55000,"Taxable","","ETF portfolio"),
     ("Home (Primary)","Real Estate","Assets",1,450000,450000,"Illiquid","","Estimated market value"),
     ("Car (Primary)","Vehicle","Assets",1,18500,18500,"Depreciating","","KBB value"),
     ("Mortgage Balance","Liability","Liabilities",-1,310000,310000,"Fixed","","30yr fixed @ 3.75%"),
     ("Car Loan","Liability","Liabilities",-1,15000,15000,"Fixed","","3yr @ 4.9%"),
     ("Student Loan","Liability","Liabilities",-1,32000,32000,"Fixed","","Federal @ 5.5%"),
     ("Credit Card Balances","Liability","Liabilities",-1,2500,2500,"Variable","","Paid monthly")],
    "Comprehensive net worth statement that tracks all assets and liabilities in one place. Automatically calculates total assets, total liabilities, and net worth. Monitor your financial health over time by updating values monthly or quarterly. Essential tool for tracking progress toward financial independence.",
    [("Asset Inventory",["List all assets with current market values",
        "Categorize by type (Cash, Retirement, Investment, Real Estate, etc.)",
        "Update values quarterly for investments, annually for real estate"]),
     ("Liability Tracking",["Record all debts with current outstanding balances",
        "Note interest rates and loan terms in the Notes column",
        "Update balances as you make payments to track debt reduction"])],
    ["Track your net worth quarterly — the trend matters more than the number",
     "Aim for net worth growth equal to or exceeding your annual income",
     "Include the cash value of life insurance and pension plans if applicable",
     "Don't include depreciating personal property (furniture, electronics)"],
    ["Net Worth = Total Assets − Total Liabilities","Liquid Net Worth = Liquid Assets − Total Liabilities",
     "Asset Allocation % = Category / Total Assets"],
    financial_keywords("Net Worth Statement"))

add_template(C1, "06_Crypto_Portfolio", "Cryptocurrency Portfolio Tracker",
    "Multi-exchange crypto portfolio with real-time P/L, allocation analysis, staking rewards & tax lot tracking",
    "portfolio",
    [("BTC","Bitcoin","2023-06-15",0.25,28500,68500,"Layer 1","HODL","Cold wallet storage"),
     ("ETH","Ethereum","2023-08-20",3.5,1800,4200,"Layer 1","Staking","Staked on Lido"),
     ("SOL","Solana","2024-01-10",50,95,175,"Layer 1","Growth","Ecosystem play"),
     ("LINK","Chainlink","2024-03-05",200,14,18.50,"Oracle","Growth","DeFi infrastructure"),
     ("AVAX","Avalanche","2024-05-15",40,35,42,"Layer 1","Growth","Subnet narrative"),
     ("MATIC","Polygon","2024-02-20",500,0.85,1.15,"L2 Scaling","Growth","ZK rollup migration"),
     ("USDC","USD Coin","2024-07-01",5000,1.00,1.00,"Stablecoin","Yield","Lending @ 8% APY"),
     ("DOT","Polkadot","2024-04-12",80,7.20,9.50,"Interop","HODL","Parachain ecosystem")],
    "Advanced cryptocurrency portfolio tracker for managing holdings across multiple exchanges and wallets. Track cost basis, current value, unrealized gain/loss, and portfolio allocation. Includes staking rewards tracking and stablecoin yield monitoring. Essential for serious crypto investors managing diversified digital asset portfolios.",
    [("Enter Holdings",["List each crypto asset with ticker, name, and purchase date",
        "Enter quantity owned and average buy price for accurate cost basis",
        "Update current prices regularly for real-time portfolio valuation"]),
     ("Portfolio Analysis",["Monitor Allocation% to ensure proper diversification",
        "Track staking rewards and yield income in the Notes column",
        "Use Strategy tags (HODL, Staking, Growth, Yield) to organize by purpose"])],
    ["Never share your actual holdings publicly — this template is for personal use only",
     "Track tokens across all wallets and exchanges in one unified view",
     "Consider tax implications before rebalancing — crypto is treated as property",
     "Keep hardware wallet holdings listed separately for security awareness"],
    ["Gain/Loss = (Current − Buy) × Quantity","Allocation% = Position / Portfolio Total",
     "Yield APY tracking in notes for passive income positions"],
    financial_keywords("Cryptocurrency Portfolio Tracker"))

# Continue with remaining finance templates (7-20) using appropriate generators
fin_remaining = [
    ("07_Real_Estate_Analyzer","Real Estate Investment Analyzer","Evaluate rental properties with cash flow, cap rate, cash-on-cash return, IRR & appreciation projections",
     "calculator",{"inputs":[("Purchase Price",350000,"$",""),("Down Payment",70000,"$","20%"),("Interest Rate",6.5,"%",""),
     ("Loan Term",30,"years",""),("Monthly Rent",2800,"$/mo",""),("Monthly Expenses",800,"$/mo","Tax+Ins+Maintenance"),
     ("Vacancy Rate",5,"%",""),("Annual Appreciation",3,"%","")],
     "calculations":[("Monthly Mortgage","=PMT(B3/12,B4*12,-(B1-B2))","$/mo",""),
     ("Gross Monthly Income","=B6","$/mo",""),("Net Operating Income","=(B6*(1-B8/100)-B7)*12","$/yr",""),
     ("Cash Flow","=(B6*(1-B8/100)-B7-B11)","$/mo",""),("Cap Rate","=B13/B1","%",""),("Cash-on-Cash Return","=(B14*12)/B2","%","")],
     "result_label":"Monthly Cash Flow"}),

    ("08_Dividend_Income","Dividend Income Tracker","Track dividend payments, yield on cost, DRIP shares & build long-term passive income streams",
     "portfolio",[("JNJ","Johnson & Johnson","2022-03-15",100,155,162,4.80,"Quarterly","3.10%","Healthcare"),
     ("PG","Procter & Gamble","2022-06-01",80,140,165,3.76,"Quarterly","2.68%","Consumer"),
     ("O","Realty Income","2023-01-10",150,58,62,3.12,"Monthly","5.38%","REIT"),
     ("KO","Coca-Cola","2022-09-20",120,60,64,1.84,"Quarterly","3.07%","Consumer")]),

    ("09_Annual_Expense","Annual Expense Tracker","12-month expense tracker with monthly category breakdowns, trend analysis & year-over-year comparison",
     "budget",{"income":[("Salary",60000,60000,"Fixed",""),("Bonus",8000,7500,"Variable","")],
     "expenses":[("Housing",18000,18000,"Fixed",""),("Food",6000,5800,"Variable",""),("Transport",2400,2200,"Variable",""),
     ("Healthcare",3600,3500,"Variable",""),("Entertainment",3600,3800,"Variable",""),("Shopping",4800,5100,"Variable","Overspent")]})
]

for fn, ttl, sub, gk, gd in fin_remaining:
    add_template(C1, fn, ttl, sub, gk, gd,
        f"Professional {ttl.lower()} to optimize your financial decisions with data-driven analysis.",
        [("Setup",[f"Enter your data in the {ttl} template","Review auto-calculated metrics","Adjust inputs to explore scenarios"]),
         ("Analysis",["Monitor trends and patterns","Compare against benchmarks","Make informed decisions based on results"])],
        ["Update data regularly for accurate tracking","Review monthly to stay on track","Share insights with financial advisors"],
        ["SUM, AVERAGE, IF for analysis","NETWORKDAYS for date calculations","PMT/FV for financial functions"],
        financial_keywords(ttl))

# Quick batch: fill remaining 11-20 with domain-appropriate generators
fin_batch = [
    ("10_Financial_Goals","Financial Goals & Milestones Planner","Set, track and achieve SMART financial goals with milestone tracking","planner",
     [("Emergency Fund","Savings",15000,8000,"2024-01-01","2025-06-30","Critical","In Progress","Self","3-6 months expenses"),
      ("Max Out Roth IRA","Retirement",7000,4000,"2024-01-01","2024-12-31","High","In Progress","Self","Tax-free growth"),
      ("Pay Off Car Loan","Debt",15000,12000,"2024-03-01","2025-03-01","High","In Progress","Self","4.9% interest"),
      ("Down Payment Fund","Savings",50000,20000,"2024-01-01","2026-12-31","Medium","In Progress","Joint","House fund")]),
    ("11_Credit_Card_Rewards","Credit Card Rewards Optimizer","Maximize credit card rewards by tracking categories, signup bonuses & optimal card usage strategy",
     "planner",[("Chase Sapphire","Travel",95000,75000,"2024-01-01","2024-12-31","High","Active","3x dining, 2x travel"),
      ("Amex Gold","Dining",80000,55000,"2024-03-01","2024-12-31","High","Active","4x dining, groceries"),
      ("Citi Double Cash","Cash Back",35000,28000,"2024-01-01","2024-12-31","Medium","Active","2% everything")]),
    ("12_Tax_Deduction_Log","Tax Deduction & Expense Log","Year-round tax deduction tracking to maximize refund and simplify filing",
     "planner",[("Home Office","Business",5000,4500,"2024-01-01","2024-12-31","High","Tracking","Deductible"),
      ("Charitable Donations","Charity",3200,2800,"2024-01-01","2024-12-31","Medium","Tracking","Receipts saved"),
      ("Professional Development","Education",1800,1500,"2024-03-01","2024-12-31","Medium","Tracking","Certifications")]),
    ("13_College_Savings_529","College Savings 529 Plan Calculator","Project education costs with inflation, 529 contribution planning & financial aid estimation",
     "calculator",{"inputs":[("Child's Current Age",8,"years",""),("College Start Age",18,"years",""),
     ("Current 529 Balance",25000,"$",""),("Annual Cost Today",25000,"$/yr",""),("Education Inflation",5,"%",""),
     ("Monthly Contribution",500,"$/mo",""),("Expected Return",6,"%",""),("Years in College",4,"years","")],
     "calculations":[("Years Until College","=B2-B1","years",""),("Future Annual Cost",f"=B4*(1+B5/100)^(B2-B1)","$/yr",""),
     ("Total Cost (4yr)",f"=B4*(1+B5/100)^(B2-B1)*B8","$",""),("529 Balance at Start",f"=FV(B7/12,(B2-B1)*12,-B6,-B3)","$",""),
     ("Shortfall/Surplus","=B10-B9","$","")],"result_label":"Funding Gap"}),
    ("14_Fee_Analyzer","Investment Fee & Expense Ratio Analyzer","Compare expense ratios across funds and visualize how fees silently erode long-term returns",
     "planner",[("VTI (0.03%)","ETF",100000,110000,"2024-01-01","2034-01-01","High","Analyzing",""),
      ("Managed Fund A (1.2%)","Mutual Fund",100000,108000,"2024-01-01","2034-01-01","High","Analyzing",""),
      ("Robo-Advisor (0.25%)","Managed",50000,52500,"2024-01-01","2034-01-01","Medium","Analyzing","")]),
    ("15_Emergency_Fund","Emergency Fund Savings Calculator","Calculate exactly how much emergency fund you need based on essential monthly expenses",
     "calculator",{"inputs":[("Monthly Essential Expenses",3500,"$/mo","Rent, food, utilities, insurance"),
     ("Current Emergency Savings",5000,"$",""),("Target Months",6,"months","3-6 recommended"),
     ("Monthly Savings Capacity",800,"$/mo",""),("Interest Rate on Savings",4,"%","HYSA rate")],
     "calculations":[("Target Fund Size","=B1*B3","$",""),("Current Shortfall","=B5-B2","$",""),
     ("Months to Reach Goal","=NPER(B4/12,-B4,B2,-B5)/12","months",""),("Date Goal Reached",f"=TODAY()+B8*30","","Estimated")],
     "result_label":"Target Fund Size"}),
    ("16_Asset_Allocation","Asset Allocation & Rebalancing Tool","Design optimal portfolio allocation with automatic rebalancing trade calculations",
     "planner",[("US Stocks","Equity",40,100000,110000,"2024-01-01","2024-12-31","High","Out of Balance","VTI, VOO"),
      ("International Stocks","Equity",25,62500,68000,"2024-01-01","2024-12-31","Medium","Out of Balance","VXUS"),
      ("Bonds","Fixed Income",20,50000,48000,"2024-01-01","2024-12-31","Medium","Out of Balance","BND"),
      ("Real Estate","Alternative",10,25000,24000,"2024-01-01","2024-12-31","Low","Out of Balance","VNQ"),
      ("Cash","Cash",5,12500,13000,"2024-01-01","2024-12-31","Low","Out of Balance","HYSA")]),
    ("17_Roth_vs_Traditional","Roth vs Traditional IRA/401(k) Comparison","Compare retirement accounts to optimize your lifetime tax strategy",
     "calculator",{"inputs":[("Current Annual Income",85000,"$",""),("Current Tax Rate",24,"%",""),
     ("Expected Retirement Tax Rate",22,"%",""),("Annual Contribution",7000,"$",""),
     ("Years Until Retirement",30,"years",""),("Expected Annual Return",7,"%","")],
     "calculations":[("Roth: After-Tax Contribution","=B4*(1-B2/100)","$/yr",""),("Roth: Future Value",f"=FV(B6/100,B5,-B4*(1-B2/100),0)","$","Tax-free"),
     ("Traditional: Future Value",f"=FV(B6/100,B5,-B4,0)","$","Pre-tax"),("Traditional: After-Tax Value","=B8*(1-B3/100)","$",""),
     ("Roth Advantage","=B7-B9","$","")],"result_label":"Roth vs Traditional Advantage"}),
    ("18_Side_Hustle","Side Hustle Income & Expense Tracker","Track multiple income streams with automatic tax set-aside calculation",
     "planner",[("Freelance Writing","Creative",2500,2300,"2024-01-01","2024-12-31","High","Active","30% tax set-aside"),
      ("Rideshare Driving","Service",1800,1900,"2024-03-01","2024-12-31","Medium","Active","25% tax set-aside"),
      ("Online Course Sales","Digital",1200,800,"2024-06-01","2024-12-31","Medium","Active","15% platform fee"),
      ("Consulting","Professional",3500,3200,"2024-01-01","2024-12-31","High","Active","35% tax set-aside")]),
    ("19_Inflation_Impact","Inflation Impact & Purchasing Power Calculator","Understand how inflation silently erodes your purchasing power over decades",
     "calculator",{"inputs":[("Current Savings",100000,"$",""),("Annual Expenses Today",50000,"$/yr",""),
     ("Inflation Rate",3,"%",""),("Time Horizon",25,"years","")],
     "calculations":[("Future Value (Nominal)",f"=B1*(1+0.05)^B4","$","Assuming 5% nominal return"),
     ("Future Expenses",f"=B2*(1+B3/100)^B4","$/yr",""),("Purchasing Power (Today's $)",f"=B6/(1+B3/100)^B4","$",""),
     ("Years Savings Will Last","=B7/B8","years",""),("Real Return","=5-B3","%","Nominal return minus inflation")],
     "result_label":"Years Savings Will Last"}),
    ("20_Bond_CD_Ladder","Bond & CD Ladder Builder","Build fixed-income bond/CD ladders for predictable, steady cash flows",
     "planner",[("1-Year CD","Fixed Income",5000,5250,"2024-06-01","2025-06-01","High","Active","5.25% APY"),
      ("2-Year Treasury","Fixed Income",5000,5400,"2024-06-01","2026-06-01","High","Active","4.85% yield"),
      ("3-Year CD","Fixed Income",5000,5600,"2024-06-01","2027-06-01","Medium","Active","4.50% APY"),
      ("5-Year Bond","Fixed Income",5000,6100,"2024-06-01","2029-06-01","Medium","Active","4.20% coupon"),
      ("iBonds","Fixed Income",5000,5350,"2024-06-01","2025-12-01","High","Active","Inflation-adjusted")]),
]
for fn, ttl, sub, gk, gd in fin_batch:
    add_template(C1, fn, ttl, sub, gk, gd,
        f"A comprehensive {ttl.lower()} designed for professional-grade financial planning and analysis.",
        [("Setup",["Enter all relevant data in the template fields","Double-check inputs for accuracy","Review auto-calculated results"]),
         ("Optimization",["Run what-if scenarios by adjusting key variables","Compare alternatives side by side","Implement the optimal strategy based on data"])],
        ["Keep records updated for accurate tracking","Review periodically (monthly or quarterly)","Use alongside professional financial advice when needed"],
        ["SUM/AVERAGE for aggregation","IF for conditional logic","Financial functions (FV, PMT, NPER, RATE) for projections"],
        financial_keywords(ttl))

# ═══════════════════════════ CATEGORY 02: PERSONAL PLANNING (20) ═══════════════════════════
C2 = "02_Personal_Planning"

plan_templates_data = [
    ("01_Annual_Goals","Annual Goal Planner & Vision Board","Set SMART yearly goals across all life dimensions with quarterly milestones and progress visualization","planner",
     [("Get promotion to Senior","Career",1,0.85,"2025-01-01","2025-12-31","Critical","In Progress","Self","Q3 review scheduled"),
      ("Run a marathon","Health",1,0.60,"2025-01-01","2025-10-15","High","In Progress","Self","Training at 15mi"),
      ("Save $30,000","Finance",30000,18500,"2025-01-01","2025-12-31","High","In Progress","Self","On track"),
      ("Read 24 books","Learning",24,14,"2025-01-01","2025-12-31","Medium","In Progress","Self","2/mo pace"),
      ("Learn Spanish A2","Learning",1,0.40,"2025-03-01","2025-12-31","Medium","In Progress","Self","Duolingo daily"),
      ("Visit Japan","Travel",1,0.90,"2025-09-01","2025-11-30","Medium","In Progress","Self","Flights booked")]),
    ("02_Daily_Routine","Daily Routine & Habit Tracker","Design your ideal daily schedule with habit consistency tracking and streak analytics","planner",
     [("Morning Exercise","Health",7,5,"2025-01-01","2025-12-31","Critical","Active","30 min workout"),
      ("Meditation","Mindfulness",7,6,"2025-01-01","2025-12-31","High","Active","15 min daily"),
      ("Reading","Learning",7,5,"2025-01-01","2025-12-31","High","Active","30 min before bed"),
      ("No Social Media AM","Focus",7,4,"2025-04-01","2025-12-31","Medium","Active","First 2 hours"),
      ("Journal Writing","Reflection",7,3,"2025-03-01","2025-12-31","Medium","Building","Evening reflection")]),
    ("03_Wheel_of_Life","Wheel of Life Balance Assessment","Assess satisfaction across 8 life dimensions and create targeted improvement action plans",
     "planner",[("Career & Work","Career",10,7,"2025-01-01","2025-06-30","Critical","Improving","Promotion goal"),
      ("Finance & Money","Finance",10,7,"2025-01-01","2025-12-31","High","Improving","Budget on track"),
      ("Health & Fitness","Health",10,6,"2025-01-01","2025-12-31","High","Improving","+3 from last year"),
      ("Family & Friends","Relationships",10,8,"2025-01-01","2025-12-31","High","Good","Monthly dinners"),
      ("Romance & Partner","Relationships",10,8,"2025-01-01","2025-12-31","High","Good","Date nights weekly"),
      ("Personal Growth","Learning",10,6,"2025-01-01","2025-12-31","Medium","Improving","Books & courses"),
      ("Fun & Recreation","Leisure",10,5,"2025-01-01","2025-08-31","Medium","Needs Work","Schedule fun time"),
      ("Spiritual & Meaning","Spirituality",10,7,"2025-01-01","2025-12-31","Medium","Good","Volunteering monthly")]),
    ("04_Weekly_Meal","Weekly Meal Planner & Grocery List","Plan nutritious meals with automated grocery list generation, calorie tracking & macro breakdown",
     "planner",[("Monday Breakfast","Breakfast",400,3.50,"2025-06-01","","High","Planned","Oatmeal + banana"),
      ("Monday Lunch","Lunch",600,8.00,"2025-06-01","","High","Planned","Chicken salad"),
      ("Monday Dinner","Dinner",700,12.00,"2025-06-01","","High","Planned","Salmon + veggies"),
      ("Tuesday Breakfast","Breakfast",380,3.00,"2025-06-02","","High","Planned","Greek yogurt + granola")]),
    ("05_Reading_List","Reading List & Book Tracker","Track books to read, reading progress, ratings, key takeaways & annual reading goals",
     "planner",[("Atomic Habits","Self-Help","2025-01-05","2025-01-25",320,"Complete","★★★★★","Systems > goals","James Clear"),
      ("Deep Work","Productivity","2025-02-01","2025-02-20",304,"Complete","★★★★★","Focus is superpower","Cal Newport"),
      ("The Psychology of Money","Finance","2025-03-01","2025-03-18",256,"Complete","★★★★★","Behavior > math","Morgan Housel"),
      ("Dune","Sci-Fi","2025-04-01","",688,"Reading","","","Frank Herbert"),
      ("Thinking Fast and Slow","Psychology","2025-05-15","",512,"To Read","","","Daniel Kahneman")]),
]

for fn, ttl, sub, gk, gd in plan_templates_data:
    add_template(C2, fn, ttl, sub, gk, gd,
        f"A comprehensive {ttl.lower()} to help you plan, track and achieve your personal goals with clarity and consistency.",
        [("Setup",["Define your goals and key metrics","Enter current status and target values","Set realistic deadlines"]),
         ("Tracking",["Update progress weekly or monthly","Review trends and adjust strategies","Celebrate milestones and wins"])],
        ["Be specific and measurable with your goals","Break large goals into smaller weekly actions","Review and reflect quarterly for course correction"],
        ["Progress% = Current / Target","AVERAGE for trends","COUNTIF for completion tracking"],
        f"{ttl.lower()},personal planning,goals,productivity,self improvement,life design,excel template,PlanoraNest")

# Fill remaining personal planning (6-20)
plan_batch = [
    ("06_Miracle_Morning","Miracle Morning Routine Planner","Design a powerful morning routine using SAVERS methodology for transformative days"),
    ("07_Bucket_List","Yearly Bucket List & Experience Tracker","Create and track your ultimate bucket list with experience categories and completion tracking"),
    ("08_Gratitude_Journal","Daily Gratitude & Positivity Journal","Cultivate gratitude with structured daily journaling prompts and mood analytics"),
    ("09_Skill_Development","Skill Development & Learning Roadmap","Plan and track deliberate skill acquisition with structured learning paths"),
    ("10_Personal_Brand","Personal Brand & Career Development","Build your professional brand with LinkedIn optimization and networking tracking"),
    ("11_Meditation_Tracker","Meditation & Mindfulness Practice Tracker","Track sessions, techniques, streaks and progress toward mindfulness mastery"),
    ("12_Digital_Detox","Digital Detox & Screen Time Tracker","Reclaim your time by tracking and systematically reducing screen time"),
    ("13_Networking_CRM","Personal Networking & Relationship CRM","Manage your professional network with a personal relationship management system"),
    ("14_Decision_Matrix","Decision Matrix & Pros/Cons Analyzer","Make better decisions with weighted criteria scoring and trade-off analysis"),
    ("15_Vision_5Year","5-Year Vision & Life Design Planner","Design your ideal life across all dimensions with backward planning from your vision"),
    ("16_30Day_Challenge","30-Day Challenge Tracker","Track daily progress for any 30-day personal challenge with streak motivation"),
    ("17_Sleep_Tracker","Sleep Quality & Improvement Tracker","Track sleep patterns, quality metrics and factors affecting nightly recovery"),
    ("18_Relationship_Plan","Relationship Investment & Connection Planner","Nurture important relationships with intentional connection and quality time planning"),
    ("19_Life_Timeline","Personal Life Timeline & Milestone Map","Map past achievements and future aspirations on your life's journey timeline"),
    ("20_Personal_SWOT","Personal SWOT Analysis & Growth Plan","Conduct a personal SWOT analysis with actionable growth and development plans"),
]
for fn, ttl, sub in plan_batch:
    add_template(C2, fn, ttl, sub, "planner",
     [("Goal 1","Personal",10000,7500,"2025-01-01","2025-12-31","High","In Progress","Self",""),
      ("Goal 2","Career",5000,3200,"2025-03-01","2025-09-30","Medium","In Progress","Self",""),
      ("Goal 3","Health",100,65,"2025-01-01","2025-06-30","High","In Progress","Self",""),
      ("Goal 4","Finance",15000,11000,"2025-02-01","2025-12-31","Medium","In Progress","Self",""),
      ("Goal 5","Learning",12,8,"2025-01-01","2025-12-31","Low","In Progress","Self","")],
     f"Professional {ttl.lower()} to help you design, track and achieve what matters most.",
     [("Design",["Clarify your vision and set specific goals","Break each goal into actionable steps","Assign priorities and realistic deadlines"]),
      ("Execute",["Track progress consistently","Review weekly and adjust monthly","Celebrate progress, not just completion"])],
     ["Start with your most important goal first","Use quarterly reviews to stay on track","Share goals with an accountability partner"],
     ["Progress = Current/Target","COUNTIF for completion","AVERAGE for trends"],
     f"{ttl.lower()},personal planning,goals,self improvement,productivity,life design,excel,PlanoraNest")

# ═══════════════════════════ CATEGORY 03: BUSINESS & MANAGEMENT (20) ═══════════════════════════
C3 = "03_Business_Management"

# Project Timeline - keep the complex one already generated
# For the rest, use appropriate generators

biz_templates_data = [
    ("01_Business_Canvas","Business Plan & Lean Canvas","One-page business plan using Lean Canvas methodology with problem-solution validation and traction metrics",
     "planner",[("Problem Validation","Strategy",10,8,"2025-01-01","2025-06-30","Critical","Validated","Customer interviews done"),
      ("Solution MVP","Product",10,7,"2025-03-01","2025-08-31","Critical","In Progress","Wireframes ready"),
      ("Customer Segments","Market",5,4,"2025-01-01","2025-05-31","High","Defined","3 personas identified"),
      ("Revenue Model","Finance",5,4,"2025-02-01","2025-07-31","High","In Progress","SaaS subscription"),
      ("Key Partnerships","Strategy",8,5,"2025-04-01","2025-09-30","Medium","In Progress","3 LOIs signed")]),
    ("02_SWOT_Analysis","Business SWOT Analysis & Strategy","Comprehensive SWOT with weighted scoring, strategic implications & action planning",
     "planner",[("Strong Brand Recognition","Strengths",9,"Internal","2025-01-01","","Critical","Leveraging","Competitive advantage"),
      ("Limited Online Presence","Weaknesses",5,"Internal","2025-01-01","2025-12-31","High","Improving","Digital transformation"),
      ("AI Market Growth","Opportunities",8,"External","2025-03-01","","High","Pursuing","$200B market by 2030"),
      ("New Entrant Competition","Threats",6,"External","2025-01-01","","High","Monitoring","3 funded startups"),
      ("Experienced Team","Strengths",8,"Internal","2025-01-01","","Medium","Leveraging","Avg 12yr experience")]),
    ("04_Cash_Flow_Forecast","Cash Flow Forecast (13-Week)","13-week rolling cash flow forecast with automatic running balance and liquidity alerts",
     "budget",{"income":[("Product Sales",25000,24000,"Recurring",""),("Service Revenue",15000,13000,"Variable",""),
     ("Investment Income",2000,1800,"Passive","")],
     "expenses":[("Payroll",15000,15000,"Fixed","Bi-weekly"),("Rent",3500,3500,"Fixed","Office"),("Marketing",5000,4800,"Variable",""),
     ("Software/Tools",2000,2100,"Variable",""),("Professional Services",3000,2500,"Variable","Legal/Accounting"),
     ("Inventory",8000,7500,"Variable",""),("Utilities",1200,1150,"Fixed",""),("Insurance",1800,1800,"Fixed","Quarterly")]}),
    ("05_Employee_Performance","Employee Performance & OKR Tracker","Track employee goals with OKRs, 360 feedback integration and development plans",
     "planner",[("Increase MRR by 25%","Business",25,18,"2025-Q1","2025-Q4","Critical","On Track","Revenue team","Linked to bonus"),
      ("Reduce Churn to <3%","Business",3,3.5,"2025-Q1","2025-Q4","Critical","At Risk","CS team","Need intervention"),
      ("Launch Mobile App","Product",1,0.7,"2025-Q2","2025-Q4","High","On Track","Engineering","Beta in Q3"),
      ("Achieve SOC2 Type II","Compliance",1,0.3,"2025-Q1","2025-Q3","High","Delayed","Security","Auditor engaged")]),
    ("06_Content_Calendar","Content Marketing Calendar","Plan multi-channel content strategy with SEO optimization, publishing schedule & performance tracking",
     "planner",[("10 SEO Tips Blog Post","Blog","2025-06-15","2025-06-20","High","Published","1200 views","SEO team"),
      ("Product Demo Video","YouTube","2025-06-22","2025-06-28","High","In Review","","Marketing"),
      ("Customer Case Study","Blog","2025-07-01","2025-07-10","Medium","Drafting","","CS team"),
      ("LinkedIn Thought Leadership","Social","2025-06-18","2025-06-18","Medium","Published","4500 impressions","CEO"),
      ("Email Newsletter #42","Email","2025-06-20","2025-06-20","High","Scheduled","","Growth team"),
      ("Industry Report","Whitepaper","2025-08-01","2025-09-15","Critical","Planning","","Research")]),
]

for fn, ttl, sub, gk, gd in biz_templates_data:
    add_template(C3, fn, ttl, sub, gk, gd,
        f"Professional {ttl.lower()} for strategic business planning and execution.",
        [("Setup",["Define your objectives and key metrics","Enter current data and baselines","Set targets and deadlines"]),
         ("Analysis",["Review calculated metrics and trends","Identify gaps and opportunities","Adjust strategy based on data insights"])],
        ["Update regularly for accurate tracking","Share with stakeholders for alignment","Review quarterly for strategic adjustments"],
        ["SUM/AVERAGE for aggregation","IF for conditional analysis","VLOOKUP for cross-referencing"],
        f"{ttl.lower()},business management,strategy,planning,excel template,professional,project management,PlanoraNest")

# Business remaining (7-20)
biz_batch = [
    ("07_Competitor_Analysis","Competitor Analysis & Market Intel","Systematic competitor tracking across product, pricing, marketing & technology dimensions"),
    ("08_SaaS_Metrics","SaaS Metrics & KPI Dashboard","Track MRR, churn, LTV, CAC, ARPU and all key SaaS growth indicators"),
    ("09_Product_Roadmap","Product Roadmap & Feature Prioritization","RICE-prioritized product roadmap with quarterly theme planning"),
    ("10_Meeting_Minutes","Meeting Agenda & Minutes Template","Professional meeting minutes with action items, decisions & follow-up tracking"),
    ("11_Expense_Report","Business Expense Report & Reimbursement","Employee expense reporting with receipt tracking and approval workflow"),
    ("12_Invoice_Tracker","Invoice Tracker & Accounts Receivable","Track invoices, payments and AR aging with automatic status updates"),
    ("13_Sales_Pipeline","Sales Pipeline & Deal Tracker","Track deals through pipeline stages with win probability and revenue forecasting"),
    ("14_Vendor_Comparison","Vendor Comparison & Selection Matrix","Weighted multi-criteria vendor evaluation and selection tool"),
    ("15_HR_Onboarding","Employee Onboarding Checklist","Comprehensive new hire onboarding with 90-day milestone tracking"),
    ("16_Risk_Register","Project Risk Register & Mitigation Tracker","Identify, assess and track project risks with probability-impact matrix"),
    ("17_Department_Budget","Department Annual Budget Planner","Department-level budget with monthly tracking, variance analysis & reforecasting"),
    ("18_Customer_Feedback","Customer Feedback & Feature Request Log","Collect, categorize and prioritize customer feedback for product development"),
    ("19_Sprint_Planning","Agile Sprint Planning & Retrospective","Sprint board with capacity planning, story points & velocity tracking"),
    ("20_Pitch_Deck_Outline","Investor Pitch Deck Outline & Fundraising","Structure your pitch deck and track investor conversations systematically"),
]
for fn, ttl, sub in biz_batch:
    add_template(C3, fn, ttl, sub, "planner",
     [("Initiative 1","Strategy",100,75,"2025-Q1","2025-Q4","High","In Progress","Owner",""),
      ("Initiative 2","Operations",80,50,"2025-Q1","2025-Q3","High","In Progress","Owner",""),
      ("Initiative 3","Growth",60,30,"2025-Q2","2025-Q4","Medium","In Progress","Owner",""),
      ("Initiative 4","Finance",50,40,"2025-Q1","2025-Q2","Medium","In Progress","Owner",""),
      ("Initiative 5","People",40,25,"2025-Q2","2025-Q4","Low","Planning","Owner","")],
     f"Professional {ttl.lower()} designed for modern business management and operational excellence.",
     [("Setup",["Define your objectives and scope","Enter current data and metrics","Set targets and timelines"]),
      ("Management",["Track progress against goals","Review metrics regularly","Adjust based on performance data"])],
     ["Keep data current for accurate insights","Use filters to focus on priorities","Share with team for alignment"],
     ["SUM/AVERAGE for totals","IF/COUNTIF for status tracking","VLOOKUP for cross-referencing"],
     f"{ttl.lower()},business,management,excel template,professional,productivity,planning,PlanoraNest")

# ═══════════════════════════ CATEGORIES 04-12: BATCH GENERATION ═══════════════════════════
# Each category gets 20 templates with appropriate domain generators

def build_category_templates(cat_dir, cat_name, templates_list, gen_type="tracker"):
    """Build a full category of 20 templates with domain-appropriate data."""
    for fn, ttl, sub in templates_list:
        if gen_type == "health":
            data = [(f"2025-0{m}-{d:02d}", f"Activity {m}", 45, 350, 5.2, 135, 72.5, 7, "Cardio", "Moderate", f"Session {m}-{d}")
                    for m, d in [(6,1),(6,3),(6,5),(6,7),(6,9),(6,11),(6,13)]]
            add_template(cat_dir, fn, ttl, sub, "health", data,
                f"Professional {ttl.lower()} for tracking and improving your health journey.",
                [("Log",["Record daily activities and health metrics","Track key indicators consistently","Monitor trends over time"]),
                 ("Optimize",["Review weekly summaries","Identify patterns and correlations","Adjust routines based on data"])],
                ["Consistency is more important than perfection","Track both quantitative and qualitative data","Share with healthcare providers when relevant"],
                ["SUM for totals","AVERAGE for trends","COUNTIF for frequency analysis"],
                f"{ttl.lower()},health,fitness,wellness,excel template,personal tracking,PlanoraNest")
        elif gen_type == "edu":
            data = [("Advanced Python Programming","Coursera","Technical","2025-06-01","2025-08-31","",75,"A-","In Progress","4 credits",49,"Monthly"),
                    ("Machine Learning Specialization","Stanford Online","Technical","2025-05-15","2025-09-15","",60,"B+","In Progress","Certificate",79,""),
                    ("Business Communication","LinkedIn Learning","Soft Skills","2025-07-01","2025-08-15","",40,"","Enrolled","1 credit",29.99,""),
                    ("Data Visualization Masterclass","Udemy","Technical","2025-08-01","","","","","Planned","Certificate",19.99,""),
                    ("Leadership Fundamentals","Harvard Online","Management","2025-09-01","","","","","Planned","Certificate",199,"")]
            add_template(cat_dir, fn, ttl, sub, "edu", data,
                f"Comprehensive {ttl.lower()} to organize and optimize your learning journey.",
                [("Plan",["Define your learning objectives and curriculum","Set realistic timelines and milestones","Allocate resources and budget"]),
                 ("Track",["Log study hours and progress","Record grades, scores and achievements","Review learning outcomes periodically"])],
                ["Focus on one major course at a time for depth","Schedule regular study blocks in your calendar","Apply learning through projects, not just courses"],
                ["NETWORKDAYS for duration","AVERAGE for grade calculation","SUM for total credits/costs"],
                f"{ttl.lower()},education,learning,study,excel template,student,academic,PlanoraNest")
        elif gen_type == "life":
            data = [("2025-06-01","Morning walk in park","Exercise","Self","Local park",4,0,"1 hr","Felt energized","",""),
                    ("2025-06-02","Coffee with mentor","Networking","Mentor","Downtown cafe",5,12,"2 hrs","Great career advice","Send thank you note",""),
                    ("2025-06-03","Volunteered at shelter","Volunteering","Community","Animal shelter",5,0,"3 hrs","Very fulfilling","Sign up for monthly",""),
                    ("2025-06-04","Movie night","Entertainment","Friends","AMC Theater",4,25,"3 hrs","Fun comedy movie","",""),
                    ("2025-06-05","Journal reflection","Self-Care","Self","Home",4,0,"1 hr","Good introspection","",""),
                    ("2025-06-06","Concert downtown","Entertainment","Partner","Stadium",5,120,"4 hrs","Amazing live music","",""),
                    ("2025-06-07","Deep cleaned house","Chores","Self","Home",2,25,"4 hrs","Productive day","","")]
            add_template(cat_dir, fn, ttl, sub, "life", data,
                f"Beautiful {ttl.lower()} to capture and enrich your daily experiences.",
                [("Capture",["Record daily events, activities and reflections","Track mood, energy and satisfaction levels","Note people, places and memorable moments"]),
                 ("Reflect",["Review weekly for patterns and insights","Identify what energizes and drains you","Adjust habits based on self-awareness"])],
                ["Journal consistently — even a few lines matters","Be honest with your ratings and reflections","Review monthly to see your growth trajectory"],
                ["AVERAGE for mood/rating trends","COUNTIF for category frequency","SUM for cost tracking"],
                f"{ttl.lower()},lifestyle,journal,personal tracking,wellness,excel template,life planning,PlanoraNest")
        elif gen_type == "home":
            data = [("Living Room","Cleaning","Weekly","2025-06-07","2025-06-14","","Self","Pending","Medium",0,"Vacuum, duster",""),
                    ("Kitchen","Deep Clean","Monthly","2025-05-28","2025-06-28","","Self","Pending","High",25,"Degreaser, cloths",""),
                    ("Garden","Maintenance","Weekly","2025-06-10","2025-06-17","","Self","Done","Medium",15,"Pruner, fertilizer",""),
                    ("HVAC Filter","Replace","Quarterly","2025-03-15","2025-06-15","","Self","Pending","High",35,"Filter size 20x25",""),
                    ("Garage","Organization","Monthly","2025-05-20","2025-06-20","","Family","Pending","Low",50,"Storage bins",""),
                    ("Windows","Washing","Monthly","2025-05-30","2025-06-30","","Self","Pending","Medium",10,"Glass cleaner",""),
                    ("Smoke Detectors","Check","Monthly","2025-05-01","2025-06-01","","Self","Overdue","Critical",5,"Batteries","")]
            add_template(cat_dir, fn, ttl, sub, "home", data,
                f"Organized {ttl.lower()} to keep your home running smoothly and efficiently.",
                [("Organize",["List all home areas, tasks and responsibilities","Set cleaning and maintenance frequencies","Assign tasks to family members"]),
                 ("Maintain",["Check Next Due dates weekly for upcoming tasks","Update status as tasks are completed","Track supplies needed for upcoming work"])],
                ["Break large home projects into smaller weekly tasks","Involve all family members in home maintenance","Set calendar reminders for recurring tasks"],
                ["NETWORKDAYS for scheduling","COUNTIF for completion tracking","SUM for supply budget"],
                f"{ttl.lower()},home,family,household,organization,cleaning,excel template,home management,PlanoraNest")
        elif gen_type == "travel":
            data = [("Tokyo, Japan","International","2025-09-15","2025-09-30","",4500,"","","Booked","JL-045","Packing list done",""),
                    ("Paris, France","International","2025-12-20","2025-12-28","",3200,"","","Researching","","","Holiday trip"),
                    ("New York City","Domestic","2025-07-10","2025-07-14","",1800,1650,"","Booked","UA-882","","Business trip"),
                    ("Bali, Indonesia","International","2026-02-01","2026-02-14","",3800,"","","Planning","","","Honeymoon"),
                    ("Weekend Camping","Outdoor","2025-08-02","2025-08-04","",300,280,"","Booked","","Tent, sleeping bag","")]
            add_template(cat_dir, fn, ttl, sub, "travel", data,
                f"Comprehensive {ttl.lower()} to plan and organize your adventures.",
                [("Plan",["Research and list all destinations","Set budgets and travel dates","Track booking status and references"]),
                 ("Execute",["Update actual expenses vs budget during travel","Log experiences and ratings for future reference","Save booking confirmations and important documents"])],
                ["Book flights and accommodations 2-3 months ahead for best rates","Always budget 15-20% extra for unexpected expenses","Keep digital copies of all travel documents"],
                ["NETWORKDAYS for trip duration","Budget vs Actual variance","SUM for total trip costs"],
                f"{ttl.lower()},travel,adventure,trip planner,vacation,excel template,travel planning,PlanoraNest")
        elif gen_type == "creative":
            data = [("Watercolor Landscape","Painting","In Progress","2025-06-01","","",4,4,35,"Watercolors, paper","Pinterest",""),
                    ("Handmade Scarf","Knitting","Complete","2025-03-15","2025-05-20","",3,5,28,"Merino wool, needles","YouTube tutorial","Gift for mom"),
                    ("Podcast Episode #5","Audio","In Progress","2025-06-05","","",3,4,0,"Mic, Audacity","","Interview format"),
                    ("Garden Redesign","Gardening","Planning","2025-07-01","","",3,0,120,"Plants, soil, tools","Magazine photos",""),
                    ("Photo Book 2025","Photography","Idea","","","",0,0,0,"","","Collecting photos all year"),
                    ("Wooden Bookshelf","Woodworking","In Progress","2025-05-01","","",5,5,85,"Oak wood, varnish","YouTube","")]
            add_template(cat_dir, fn, ttl, sub, "creative", data,
                f"Inspiring {ttl.lower()} to nurture your creativity and track your creative journey.",
                [("Create",["Log all creative projects and ideas","Track time spent and materials used","Rate difficulty and satisfaction"]),
                 ("Grow",["Review completed projects for learning","Plan next creative challenges","Build on skills progressively"])],
                ["Don't wait for perfection — start creating and iterate","Track your progress to see skill improvement over time","Join creative communities for feedback and inspiration"],
                ["NETWORKDAYS for project duration","AVERAGE for satisfaction trends","SUM for material costs"],
                f"{ttl.lower()},creative,hobbies,crafts,art,DIY,excel template,creativity,PlanoraNest")
        elif gen_type == "productivity":
            data = [("Write Q3 Strategy Doc","Deep Work","Critical",4,3.5,"","8",2,"Yes","2025-06-15","Pomodoro",""),
                    ("Email Inbox Zero","Admin","Medium",1,0.8,"","6",5,"Yes","2025-06-14","GTD",""),
                    ("Code Review Session","Deep Work","High",2,2.2,"","9",1,"Yes","2025-06-14","Time Blocking",""),
                    ("Team Standup Meeting","Meeting","Medium",0.5,0.5,"","7",3,"Yes","2025-06-15","","Daily"),
                    ("Read Industry Reports","Learning","Low",1.5,1.8,"","5",4,"No","","",""),
                    ("Plan Sprint Backlog","Planning","High",2,2.5,"","8",0,"Yes","2025-06-13","Deep Work",""),
                    ("Social Media Management","Shallow","Low",0.5,1.2,"","4",8,"Yes","2025-06-14","","Batch process")]
            add_template(cat_dir, fn, ttl, sub, "productivity", data,
                f"Powerful {ttl.lower()} to maximize your effectiveness and accomplish what matters.",
                [("Plan",["List all tasks and categorize by type","Assign priority levels and time estimates","Block time in your calendar for focused work"]),
                 ("Track",["Record actual time spent vs estimated","Rate focus quality and distraction count","Analyze efficiency patterns weekly"])],
                ["Schedule deep work during your peak energy hours","Batch similar tasks together to reduce context switching","The goal isn't to be busy — it's to be effective"],
                ["Efficiency% = Estimated / Actual × 100","AVERAGE for focus trends","COUNTIF for completion tracking"],
                f"{ttl.lower()},productivity,time management,focus,GTD,deep work,excel template,efficiency,PlanoraNest")
        elif gen_type == "it":
            data = [("AST-001","MacBook Pro 16\"","Hardware","Engineering","2024-06-15",2499,25,"","2027-06-15","Alice Chen","Active","HQ-3F-12",""),
                    ("AST-002","Dell UltraSharp 27\"","Hardware","Design","2024-08-20",699,20,"","2027-08-20","Bob Wang","Active","HQ-3F-12",""),
                    ("AST-003","AWS Production Account","Cloud","DevOps","2024-01-01",0,0,"","","Infra Team","Active","us-east-1","Monthly ~$4500"),
                    ("AST-004","GitHub Enterprise","Software","Engineering","2024-01-15",0,0,"","2025-01-15","All Engineers","Active","","42/user/mo"),
                    ("AST-005","Cisco Switch 9200","Network","IT","2024-03-01",3200,15,"","2027-03-01","Network Admin","Active","Server Room",""),
                    ("AST-006","Jira Cloud","Software","Product","2024-01-01",0,0,"","2025-01-01","All Teams","Active","","$15/user/mo"),
                    ("AST-007","iPhone 15 Pro","Hardware","Sales","2024-10-01",999,30,"","2026-10-01","Sales Team","Active","Field","10 devices")]
            add_template(cat_dir, fn, ttl, sub, "it", data,
                f"Enterprise-grade {ttl.lower()} for professional IT asset and infrastructure management.",
                [("Inventory",["Catalog all IT assets with unique IDs and details","Categorize by type, department, and location","Track purchase dates, costs, and warranty periods"]),
                 ("Manage",["Monitor asset status and current values","Plan refresh cycles based on depreciation","Audit software licenses for compliance"])],
                ["Conduct quarterly physical asset audits","Keep software license keys in a secure separate document","Set calendar reminders 3 months before warranty expiration"],
                ["Current Value = Cost × (1 − Depreciation%)","COUNTIF for status summary","SUM for total asset value"],
                f"{ttl.lower()},IT,technology,asset management,infrastructure,excel template,IT management,PlanoraNest")
        elif gen_type == "product_mgmt":
            data = [("User Dashboard v2","Core","","12000",5,80,"8","2025-Q3","In Progress","Auth system","PM",""),
                    ("API Rate Limiting","Infrastructure","","8500",4,70,"6","2025-Q3","In Progress","","Engineering",""),
                    ("Dark Mode Support","UX","","5000",3,90,"4","2025-Q4","Planned","","Design",""),
                    ("SSO Integration","Enterprise","","15000",4,60,"10","2025-Q4","Planned","Auth system","Engineering",""),
                    ("Export to CSV","Feature","","3000",2,95,"2","2025-Q3","In Progress","","Engineering",""),
                    ("Onboarding Wizard","Growth","","10000",5,75,"8","2025-Q3","In Progress","","PM",""),
                    ("Performance Optimization","Infrastructure","","8000",3,70,"6","2025-Q4","Planned","","Engineering","")]
            add_template(cat_dir, fn, ttl, sub, "product_mgmt", data,
                f"Professional {ttl.lower()} to build better products with structured product management.",
                [("Plan",["Define features and initiatives with clear scope","Score and prioritize using RICE or similar framework","Set quarterly themes and roadmap timelines"]),
                 ("Execute",["Track development progress and dependencies","Monitor status and remove blockers","Review metrics and impact post-launch"])],
                ["Focus on outcomes, not output — measure impact, not velocity","Say no to good ideas to make room for great ones","Involve engineering early in the prioritization process"],
                ["RICE = Reach × Impact × Confidence / Effort","COUNTIF for status tracking","SUM for effort estimation"],
                f"{ttl.lower()},product management,roadmap,features,prioritization,excel template,product development,PlanoraNest")
        else:
            # Default tracker
            data = [(f"2025-0{6+i//3}-{(i%3)*10+1:02d}", f"Entry {i+1}", f"Description for entry {i+1}", "Category A", "", "")
                    for i in range(8)]
            add_template(cat_dir, fn, ttl, sub, "tracker", data,
                f"Professional {ttl.lower()} for organizing and tracking your activities.",
                [("Setup",["Define your tracking categories and metrics","Enter initial data","Set goals and targets"]),
                 ("Maintain",["Log entries consistently","Review patterns weekly","Adjust approach as needed"])],
                ["Consistency is key to meaningful tracking","Review data monthly for insights","Share with relevant stakeholders"],
                ["SUM/AVERAGE for analysis","COUNTIF for counting","IF for conditional logic"],
                f"{ttl.lower()},{cat_name.lower()},excel template,productivity,tracking,organization,PlanoraNest")

# Define all remaining templates across categories 4-12
ALL_CATEGORIES = {
    "04_Health_Fitness": ("Health & Fitness", "health", [
        ("01_Workout_Planner","Workout & Exercise Planner","Plan workouts, track sets/reps, monitor progress & calculate 1RM"),
        ("02_Calorie_Counter","Daily Calorie & Macro Counter","Track calories, protein, carbs, fats with barcode scanning & meal breakdowns"),
        ("03_Running_Log","Running & Race Training Log","Log runs with pace, distance, heart rate, elevation & race PR tracking"),
        ("04_Yoga_Practice","Yoga Practice & Pose Tracker","Track sessions, poses mastered, meditation minutes & flexibility progress"),
        ("05_Weight_Loss","Weight Loss & Body Transformation","Track weight, measurements, body fat % with trend charts & milestone celebrations"),
        ("06_Gym_Progress","Gym Progress & Strength Tracker","Log exercises, sets, reps, 1RM calculations & progressive overload tracking"),
        ("07_Meal_Nutrition","Meal Nutrition & Diet Planner","Plan nutritious meals with macro breakdowns, grocery integration & meal prep"),
        ("08_Water_Intake","Water Intake & Hydration Tracker","Track daily water consumption with goals, reminders & health correlation"),
        ("09_Steps_Counter","Daily Steps & Activity Tracker","Track steps, active minutes, flights climbed & weekly activity summaries"),
        ("10_Fitness_Goals","Fitness Goals & PR Tracker","Set fitness goals, track personal records & celebrate achievement milestones"),
        ("11_Body_Measurements","Body Measurements & Progress Photos","Log measurements across body points with progress photo timeline integration"),
        ("12_Cycling_Log","Cycling Ride & Performance Log","Track rides with distance, elevation, power data & segment PRs"),
        ("13_HIIT_Timer","HIIT Workout Timer & Interval Planner","Plan high-intensity intervals with work/rest ratios & round tracking"),
        ("14_Stretching_Routine","Stretching & Flexibility Routine","Design stretching routines with hold times, progress photos & mobility goals"),
        ("15_Wellness_Journal","Wellness & Self-Care Journal","Daily wellness check-ins with mood, energy & self-care activity tracking"),
        ("16_Strength_Training","Strength Training Program Builder","Build progressive overload programs with exercise library & schedule"),
        ("17_Cardio_Plan","Cardio Training & Endurance Plan","Design cardio programs with heart rate zones & endurance progression"),
        ("18_Health_Screening","Health Screening & Medical Log","Track screenings, vaccinations, lab results & family health history"),
        ("19_Vitamin_Supplement","Vitamin & Supplement Tracker","Log supplements, dosages, timing & effectiveness ratings"),
        ("20_PT_Sessions","Personal Training Session Log","Track PT sessions with exercises, feedback & progress toward goals"),
    ]),
    "05_Education_Learning": ("Education & Learning", "edu", [
        ("01_Study_Planner","Study Planner & Schedule","Plan study sessions with subject rotation, breaks & exam preparation timeline"),
        ("02_Course_Catalog","Course Catalog & Curriculum Tracker","Browse and track courses across platforms with ratings & completion status"),
        ("03_Grade_Calculator","Grade & GPA Calculator","Calculate current grades, project final grades & track GPA across semesters"),
        ("04_Research_Notes","Research Notes & Literature Review","Organize research with structured notes, citations & thematic analysis"),
        ("05_Language_Learning","Language Learning Progress Tracker","Track vocabulary, grammar topics, practice hours & fluency milestones"),
        ("06_Certification_Tracker","Certification & Credential Tracker","Track certifications with renewal dates, CE credits & exam preparation"),
        ("07_Online_Courses","Online Course Progress Tracker","Track courses across Coursera, Udemy, edX with progress & completion dates"),
        ("08_Thesis_Outline","Thesis & Dissertation Planner","Structure thesis chapters, track word count, manage advisor feedback"),
        ("09_Flashcards_System","Flashcard Review & Spaced Repetition","Schedule flashcard reviews using spaced repetition intervals"),
        ("10_Exam_Prep","Exam Preparation & Study Tracker","Plan exam prep with topic coverage, practice tests & confidence ratings"),
        ("11_Assignment_Tracker","Assignment & Homework Tracker","Track assignments across classes with due dates, weight & grade tracking"),
        ("12_GPA_Calculator","GPA & Academic Performance Tracker","Comprehensive GPA calculation with credit hours & grade point tracking"),
        ("13_Study_Group","Study Group Organizer & Notes","Coordinate study groups with agenda, shared notes & action items"),
        ("14_Academic_Calendar","Academic Semester Calendar","Plan entire semester with class schedule, exam dates & project deadlines"),
        ("15_Scholarship_Tracker","Scholarship & Grant Application Tracker","Track scholarship applications with deadlines, requirements & award status"),
        ("16_Lecture_Notes","Lecture Notes & Cornell Template","Structured note-taking with Cornell method: cues, notes & summary sections"),
        ("17_Coding_Practice","Coding Practice & Algorithm Tracker","Track LeetCode/HackerRank problems with solutions, complexity & patterns"),
        ("18_Library_Books","Library Books & Reading Tracker","Track borrowed books with due dates, renewals & reading progress"),
        ("19_Project_Based_Learning","Project-Based Learning Planner","Design and track learning through hands-on projects & portfolios"),
        ("20_Degree_Planner","Degree & Graduation Path Planner","Map course requirements, prerequisites & graduation timeline"),
    ]),
    "06_Life_Lifestyle": ("Life & Lifestyle", "life", [
        ("01_Daily_Journal","Daily Reflection Journal","Capture daily thoughts, experiences & lessons with structured journaling prompts"),
        ("02_Mood_Tracker","Mood & Emotion Tracker","Log daily moods with triggers, coping strategies & emotional pattern analysis"),
        ("03_Volunteer_Log","Volunteer Hours & Impact Log","Track volunteer activities, hours contributed & community impact"),
        ("04_Pet_Care","Pet Care & Health Tracker","Manage pet vaccinations, vet visits, grooming & daily care routines"),
        ("05_Subscription_Manager","Subscription & Recurring Expense Manager","Track all subscriptions with costs, renewal dates & usage assessment"),
        ("06_Event_Planner","Event Planning & Party Organizer","Plan events with guest lists, budgets, timelines & vendor coordination"),
        ("07_Wardrobe_Inventory","Wardrobe Inventory & Outfit Planner","Catalog clothing items, plan outfits & track cost-per-wear"),
        ("08_Self_Care_Routine","Self-Care Routine & Wellness Planner","Design and track self-care activities with frequency & satisfaction ratings"),
        ("09_Bucket_List_Advanced","Ultimate Bucket List & Life Experiences","Curate life goals across adventure, learning, contribution & legacy categories"),
        ("10_Birthday_Calendar","Birthday & Important Dates Calendar","Never miss birthdays, anniversaries & special occasions with reminders"),
        ("11_Password_Manager","Password & Digital Security Manager","Track password rotations, 2FA status & security audit schedule (local only)"),
        ("12_Journal_Prompts","Journal Prompts & Writing Inspiration","Curated collection of journal prompts with writing space & theme tracking"),
        ("13_Morning_Evening","Morning & Evening Routine Designer","Design AM/PM routines with habit stacking & wind-down optimization"),
        ("14_Acts_of_Kindness","Random Acts of Kindness Tracker","Track kind acts given and received with ripple effect documentation"),
        ("15_Decluttering_Plan","Decluttering & Minimalism Plan","Room-by-room decluttering with KonMari categories & donation tracking"),
        ("16_Personal_Manifesto","Personal Manifesto & Values Statement","Craft your personal mission statement, core values & guiding principles"),
        ("17_Life_Audit","Life Audit & Satisfaction Assessment","Comprehensive life audit across all domains with gap analysis"),
        ("18_Values_Assessment","Core Values Assessment & Alignment","Identify, prioritize & align actions with your personal core values"),
        ("19_Coaching_Notes","Life Coaching Session Notes","Structured coaching session notes with goals, insights & action commitments"),
        ("20_Year_In_Review","Annual Year in Review & Reflection","Comprehensive year-end reflection with achievements, lessons & next year vision"),
    ]),
    "07_Home_Family": ("Home & Family", "home", [
        ("01_Home_Inventory","Home Inventory & Insurance Record","Document all possessions with photos, values & insurance documentation"),
        ("02_Cleaning_Schedule","House Cleaning Schedule & Checklist","Room-by-room cleaning schedule with daily/weekly/monthly task rotation"),
        ("03_Chore_Chart","Family Chore Chart & Rotation","Assign and rotate household chores with completion tracking & rewards"),
        ("04_Garden_Planner","Garden Design & Plant Care Planner","Plan garden layouts, track planting dates & monitor plant health"),
        ("05_Home_Renovation","Home Renovation Project Planner","Plan renovations with budgets, contractor bids, timeline & material lists"),
        ("06_Moving_Checklist","Moving House Checklist & Planner","Comprehensive moving checklist with timeline, packing inventory & utility transfers"),
        ("07_Room_Designer","Room Design & Furniture Layout","Design room layouts with measurements, furniture placement & shopping list"),
        ("08_Appliance_Tracker","Appliance & Warranty Tracker","Track appliances with purchase dates, warranties & maintenance schedules"),
        ("09_Mortgage_Calculator","Mortgage & Home Loan Calculator","Calculate payments, amortization & compare loan scenarios"),
        ("10_Utility_Bills","Utility Bill Tracker & Energy Monitor","Track monthly utility usage & costs with year-over-year comparison"),
        ("11_Furniture_Wishlist","Furniture & Decor Wishlist","Curate home furnishing wishlist with prices, dimensions & priority ranking"),
        ("12_Paint_Colors","Paint Color & Room Palette Planner","Plan paint colors with swatch tracking, finish types & room coordination"),
        ("13_Home_Maintenance","Seasonal Home Maintenance Schedule","Year-round maintenance checklist organized by season & priority"),
        ("14_Garage_Inventory","Garage & Storage Inventory","Organize garage, basement & attic storage with location mapping"),
        ("15_Plant_Care","Indoor Plant Care & Watering Schedule","Track indoor plants with watering, fertilizing & repotting schedules"),
        ("16_Home_Security","Home Security & Safety Checklist","Security audit, camera placement, emergency contacts & safety protocols"),
        ("17_Energy_Audit","Home Energy Audit & Efficiency Plan","Track energy usage, identify waste & plan efficiency improvements"),
        ("18_Storage_System","Storage Organization System","Design storage solutions with bin labels, locations & seasonal rotation"),
        ("19_Rental_Property","Rental Property Management","Manage rental properties with tenant info, lease dates & maintenance requests"),
        ("20_Holiday_Planner","Holiday & Celebration Planner","Plan holiday menus, decorations, gift lists & traditions"),
    ]),
    "08_Travel_Adventure": ("Travel & Adventure", "travel", [
        ("01_Trip_Planner","Comprehensive Trip Planner","Plan every detail: flights, hotels, activities, budgets & daily itinerary"),
        ("02_Packing_Checklist","Smart Packing Checklist & Organizer","Never forget essentials with weather-based, activity-specific packing lists"),
        ("03_Travel_Budget","Travel Budget & Expense Tracker","Track travel spending with currency conversion & budget vs actual analysis"),
        ("04_Flight_Tracker","Flight Price & Award Ticket Tracker","Track flight prices, award availability & frequent flyer mile optimization"),
        ("05_Road_Trip_Route","Road Trip Route & Stop Planner","Plan road trip routes with stop planning, fuel costs & scenic detours"),
        ("06_Hotel_Comparison","Hotel & Accommodation Comparison","Compare hotels across price, location, amenities & reviews"),
        ("07_Travel_Insurance","Travel Insurance Comparison","Compare travel insurance policies with coverage details & exclusions"),
        ("08_Destination_Research","Destination Research & Wishlist","Research destinations with attractions, best times to visit & local tips"),
        ("09_Solo_Travel","Solo Travel Safety & Planning","Solo-specific planning with safety tips, social opportunities & self-care"),
        ("10_Family_Vacation","Family Vacation Multi-Gen Planner","Plan family trips with activities for all ages, kid-friendly dining & logistics"),
        ("11_Business_Travel_Log","Business Travel Expense & Trip Log","Track business trips with per-diem, mileage & reimbursement tracking"),
        ("12_Frequent_Flyer","Frequent Flyer & Points Optimizer","Track loyalty programs, elite status progress & points expiration dates"),
        ("13_Backpacking_Gear","Backpacking Gear List & Weight Calculator","Pack with weight optimization, gear checklist & trail preparation"),
        ("14_Currency_Converter","Multi-Currency Travel Wallet","Track spending across currencies with real-time exchange rate conversion"),
        ("15_Local_Food_Guide","Local Food & Restaurant Discovery","Document food discoveries, local specialties & restaurant recommendations"),
        ("16_Travel_Itinerary","Daily Travel Itinerary Planner","Hour-by-hour daily itinerary with activities, transport & free time"),
        ("17_Bucket_List_Dest","Ultimate Travel Bucket List","Curate and track dream destinations with seasons, budgets & priority ranking"),
        ("18_Cruise_Planner","Cruise Vacation Planner","Plan cruise with shore excursions, dining reservations & onboard activities"),
        ("19_Camping_Checklist","Camping Trip Planner & Checklist","Plan camping trips with gear, meal planning & campsite reservations"),
        ("20_Travel_Photo_Log","Travel Photo & Memory Journal","Organize travel photos with locations, stories & print-worthy selections"),
    ]),
    "09_Creative_Hobbies": ("Creative & Hobbies", "creative", [
        ("01_Art_Project_Tracker","Art Project & Portfolio Tracker","Track art projects with progress photos, techniques used & exhibition planning"),
        ("02_Writing_Journal","Creative Writing & Story Tracker","Track writing projects with word count, character development & submissions"),
        ("03_Music_Practice","Music Practice & Repertoire Tracker","Log practice sessions, track pieces learned & prepare for performances"),
        ("04_Podcast_Planner","Podcast Episode Planner & Content Calendar","Plan podcast episodes with topics, guests, show notes & publishing schedule"),
        ("05_Photography_Log","Photography Shoot & Gear Tracker","Log shoots with settings, locations, gear used & post-processing notes"),
        ("06_DIY_Projects","DIY Project Planner & Material List","Plan DIY projects with steps, materials, tools & cost tracking"),
        ("07_Craft_Inventory","Craft Supply Inventory & Stash Organizer","Organize craft supplies with quantities, storage locations & project ideas"),
        ("08_Video_Production","Video Production & YouTube Planner","Plan videos with scripts, shot lists, editing timeline & publishing schedule"),
        ("09_Knitting_Crochet","Knitting & Crochet Project Tracker","Track patterns, yarn stash, needle sizes & project progress"),
        ("10_Woodworking_Plan","Woodworking Project & Cut List","Design woodworking projects with cut lists, joinery & finishing notes"),
        ("11_Painting_Tracker","Painting Series & Technique Journal","Track painting series with color palettes, techniques & gallery submissions"),
        ("12_Sewing_Projects","Sewing Pattern & Fabric Stash Organizer","Track patterns, fabric inventory, measurements & alteration notes"),
        ("13_Recipe_Collection","Recipe Collection & Meal Memory Book","Collect recipes with ratings, modifications & occasion pairings"),
        ("14_Garden_Design","Landscape Design & Plant Journal","Design garden spaces with plant selections, bloom calendars & care notes"),
        ("15_Calligraphy_Practice","Calligraphy & Lettering Practice Log","Track practice sessions with style progression & project applications"),
        ("16_Pottery_Tracker","Pottery & Ceramics Studio Log","Track pieces from clay to kiln with glazes, firing notes & results"),
        ("17_Digital_Art","Digital Art & Illustration Tracker","Track digital pieces with software, brushes, layers & commission details"),
        ("18_Band_Rehearsal","Band Rehearsal & Setlist Planner","Plan rehearsals, track setlists & prepare for gigs"),
        ("19_Animation_Project","Animation Project & Timeline Tracker","Track animation projects with frames, scenes & rendering progress"),
        ("20_Creative_Business","Creative Business & Etsy Shop Manager","Manage creative business with inventory, pricing & order tracking"),
    ]),
    "10_Productivity_Time": ("Productivity & Time Management", "productivity", [
        ("01_Pomodoro_Tracker","Pomodoro Session & Focus Tracker","Track Pomodoro sessions with task association & daily focus metrics"),
        ("02_Weekly_Planner","Ultimate Weekly Planner & Organizer","Plan your ideal week with time blocks, priorities & weekly review"),
        ("03_Eisenhower_Matrix","Eisenhower Priority Decision Matrix","Sort tasks by urgency & importance with automatic quadrant placement"),
        ("04_Deep_Work_Blocks","Deep Work Session Tracker & Planner","Schedule and track deep work blocks with distraction log & output metrics"),
        ("05_GTD_System","Getting Things Done Complete System","Full GTD implementation with inbox, projects, contexts & weekly review"),
        ("06_Time_Audit","Time Audit & Leak Detector","Track every hour for a week to discover where your time actually goes"),
        ("07_Focus_Sessions","Focus Session & Flow State Tracker","Track focus intensity, duration & conditions that enable flow state"),
        ("08_Priority_Matrix","Impact-Effort Priority Scoring Matrix","Score tasks by impact and effort to identify quick wins & major projects"),
        ("09_Energy_Management","Energy & Chronotype Tracker","Track energy levels by hour to align tasks with natural energy peaks"),
        ("10_Quarterly_Goals","Quarterly OKR & 90-Day Goal Planner","Set quarterly objectives with key results and weekly milestone tracking"),
        ("11_Inbox_Zero","Email Management & Inbox Zero Tracker","Track email processing with response time, batching & newsletter management"),
        ("12_Task_Batching","Task Batching & Theme Day Planner","Design themed days and batched task blocks for maximum efficiency"),
        ("13_Project_Checklist","Master Project Checklist & SOP Builder","Create standard operating procedures with checklist automation"),
        ("14_Daily_Highlights","Daily Highlight & Wins Journal","Capture daily wins, highlights & gratitude for motivation & growth"),
        ("15_Time_Blocking","Time Blocking Schedule & Calendar","Design your ideal week with color-coded time blocks & buffer zones"),
        ("16_Brain_Dump","Brain Dump & Mind Sweep Tool","Complete mind sweep with categorization & action extraction"),
        ("17_Productivity_Score","Productivity Score & Habit Dashboard","Score daily productivity with weighted habit & output metrics"),
        ("18_Weekly_Review","Weekly Review & Planning System","Comprehensive weekly review with reflect, plan & organize phases"),
        ("19_Annual_Reflection","Annual Reflection & Life Assessment","Deep annual review across all life dimensions with next-year vision"),
        ("20_Master_Task_List","Master Task List & Life Dashboard","Centralized task management with projects, areas & tags"),
    ]),
    "11_IT_Technology": ("IT & Technology", "it", [
        ("01_IT_Asset_Inventory","IT Asset Inventory & Lifecycle Manager","Complete IT asset tracking with procurement, deployment & retirement"),
        ("02_Network_Diagram","Network Topology & Device Map","Document network architecture with devices, IPs & connectivity"),
        ("03_Server_Monitoring","Server Health & Monitoring Dashboard","Track server metrics: CPU, memory, disk, uptime & incidents"),
        ("04_Software_Licenses","Software License & Compliance Manager","Track licenses with expiry, renewal costs & compliance status"),
        ("05_IT_Incident_Report","IT Incident Report & RCA Tracker","Log incidents with severity, root cause analysis & resolution tracking"),
        ("06_Change_Management","IT Change Management & Approval","Track changes with risk assessment, approval workflow & rollback plans"),
        ("07_Backup_Schedule","Backup Schedule & Recovery Test Log","Track backup jobs with success rates, retention & recovery testing"),
        ("08_Cybersecurity_Risk","Cybersecurity Risk Assessment Matrix","Assess security risks with threat modeling & mitigation planning"),
        ("09_Patch_Management","Patch Management & Update Tracker","Track patches across systems with severity, testing & deployment status"),
        ("10_IT_Budget","IT Budget & Technology Spending","Plan IT budget with hardware, software, services & cloud cost tracking"),
        ("11_Tech_Documentation","Technical Documentation & Knowledge Base","Organize technical docs with version tracking & review status"),
        ("12_API_Documentation","API Inventory & Documentation Hub","Catalog APIs with endpoints, versions, auth methods & deprecation status"),
        ("13_Database_Schema","Database Schema & Migration Tracker","Document schemas, track migrations & monitor performance"),
        ("14_Code_Review_Checklist","Code Review Checklist & Quality Gate","Structured code review with security, performance & style criteria"),
        ("15_DevOps_Pipeline","DevOps Pipeline & CI/CD Monitor","Track pipeline stages, build times, deployment frequency & failure rates"),
        ("16_Cloud_Resources","Cloud Resource & Cost Inventory","Track cloud resources across providers with cost allocation & optimization"),
        ("17_Bug_Tracker","Bug & Defect Tracking System","Track bugs with severity, reproduction steps & fix verification"),
        ("18_System_Architecture","System Architecture Decision Records","Document architecture decisions with context, options & rationale"),
        ("19_IT_Onboarding","IT Onboarding & Access Provisioning","Manage new hire IT setup with access requests & equipment assignment"),
        ("20_SLA_Tracker","SLA Performance & Uptime Tracker","Track service level agreements with uptime, response time & penalties"),
    ]),
    "12_Product_Management": ("Product Management", "product_mgmt", [
        ("01_PRD_Document","Product Requirements Document (PRD)","Structured PRD with problem statement, user stories & acceptance criteria"),
        ("02_User_Story_Map","User Story Mapping & Journey Board","Map user journeys with stories organized by activity & release"),
        ("03_Launch_Checklist","Product Launch Checklist & GTM","Comprehensive launch checklist with marketing, support & engineering readiness"),
        ("04_Feature_Flags","Feature Flag & Toggle Management","Track feature flags with rollout %, target audience & cleanup schedule"),
        ("05_AB_Test_Results","A/B Test Design & Results Analyzer","Design experiments with hypothesis, sample size & statistical significance"),
        ("06_User_Personas","User Persona & Empathy Map Builder","Document personas with demographics, goals, pains & behavior patterns"),
        ("07_Customer_Journey","Customer Journey & Touchpoint Map","Map end-to-end journey with touchpoints, emotions & opportunities"),
        ("08_Product_Analytics","Product Analytics & Event Tracking","Define analytics events, properties & track key product metrics"),
        ("09_Competitive_Features","Competitive Feature Analysis Matrix","Compare features across competitors with differentiation scoring"),
        ("10_Beta_Program","Beta Testing Program Manager","Manage beta testers, feedback collection & iteration tracking"),
        ("11_Product_OKRs","Product OKR & Goal Cascade","Set product OKRs aligned with company goals & track quarterly progress"),
        ("12_Stakeholder_Plan","Stakeholder Communication Plan","Map stakeholders with communication preferences & update cadence"),
        ("13_Pricing_Analysis","Pricing Strategy & Tier Analysis","Analyze pricing models with competitor benchmarking & willingness-to-pay"),
        ("14_Product_Sunset","Product End-of-Life & Sunset Plan","Plan EOL with migration timeline, customer communication & data export"),
        ("15_UX_Audit","UX Heuristic Evaluation & Audit","Conduct usability audit with Nielsen heuristics & severity ratings"),
        ("16_NPS_Survey","NPS Survey & Sentiment Tracker","Track Net Promoter Score with response analysis & follow-up actions"),
        ("17_Adoption_Funnel","Feature Adoption & Activation Funnel","Track discovery → activation → adoption → retention funnel metrics"),
        ("18_Release_Notes","Release Notes & Changelog Builder","Craft user-facing release notes with feature highlights & bug fixes"),
        ("19_Product_FAQ","Product FAQ & Knowledge Base Builder","Build comprehensive FAQ with categories, searchability & update tracking"),
        ("20_North_Star_Metrics","North Star Metric & KPI Tree","Define north star with input metrics, KPI tree & leading indicators"),
    ]),
}

print("=" * 60)
print("PlanoraNest 240 Professional Excel Template Generator v2.0")
print(f"Website: {SITE}")
print("=" * 60)
print(f"\nTotal template definitions: {len(ALL_TEMPLATES)}")
print("Building remaining category templates...")

# Add all category templates
for cat_dir, (cat_name, gen_type, templates) in ALL_CATEGORIES.items():
    build_category_templates(cat_dir, cat_name, templates, gen_type)
    print(f"  [OK] Added {len(templates)} templates for {cat_name}")

print(f"\nFinal total: {len(ALL_TEMPLATES)} templates defined")
print("=" * 60)

# ═══════════════════════════ GENERATION ENGINE ═══════════════════════════

def generate_template(cat_dir, filename, title, subtitle, gen_key, gen_data, overview, steps, tips, formulas, keywords):
    """Generate a single complex template."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Build the appropriate data based on generator key
    if gen_key == "portfolio":
        hdrs, wids, data, dv, cf, summary, freeze = gen_portfolio_tracker(title, gen_data)
    elif gen_key == "budget":
        hdrs, wids, data, dv, cf, summary, freeze = gen_budget_planner(["Category"], gen_data.get("income",[]), gen_data.get("expenses",[]))
    elif gen_key == "planner":
        hdrs, wids, data, dv, cf, summary, freeze = gen_planner(title, gen_data)
    elif gen_key == "calculator":
        hdrs, wids, data, dv, cf, summary, freeze = gen_calculator(title, gen_data.get("inputs",[]), gen_data.get("calculations",[]), gen_data.get("result_label","Result"))
    elif gen_key == "health":
        hdrs, wids, data, dv, cf, summary, freeze = gen_health_tracker(title, gen_data)
    elif gen_key == "edu":
        hdrs, wids, data, dv, cf, summary, freeze = gen_edu_tracker(title, gen_data)
    elif gen_key == "life":
        hdrs, wids, data, dv, cf, summary, freeze = gen_life_tracker(title, gen_data)
    elif gen_key == "home":
        hdrs, wids, data, dv, cf, summary, freeze = gen_home_tracker(title, gen_data)
    elif gen_key == "travel":
        hdrs, wids, data, dv, cf, summary, freeze = gen_travel_tracker(title, gen_data)
    elif gen_key == "creative":
        hdrs, wids, data, dv, cf, summary, freeze = gen_creative_tracker(title, gen_data)
    elif gen_key == "productivity":
        hdrs, wids, data, dv, cf, summary, freeze = gen_productivity_tracker(title, gen_data)
    elif gen_key == "it":
        hdrs, wids, data, dv, cf, summary, freeze = gen_it_asset(title, gen_data)
    elif gen_key == "product_mgmt":
        hdrs, wids, data, dv, cf, summary, freeze = gen_product_template(title, gen_data)
    elif gen_key == "tracker":
        hdrs, wids, data, dv, cf, summary, freeze = gen_tracker_log(title, gen_data)
    else:
        hdrs, wids, data, dv, cf, summary, freeze = gen_tracker_log(title, gen_data)

    # Create sheets
    ws_data = wb.create_sheet("📊 Template")
    make_template_sheet(ws_data, title, subtitle, hdrs, wids, data, dv, cf, summary, freeze)

    ws_ins = wb.create_sheet("📖 Instructions")
    make_instructions_sheet(ws_ins, title, overview, steps, tips, formulas)

    ws_abt = wb.create_sheet("ℹ️ About")
    cat_name = CATS.get(cat_dir, cat_dir.replace("_", " "))
    extra = [f"{len(hdrs)} columns with structured data entry",
             f"{len(data)} sample rows with pre-filled data",
             "Automatic formula calculations",
             "Data validation dropdowns for consistency",
             "Conditional formatting for visual insights"]
    make_about_sheet(ws_abt, title, cat_name, keywords, extra)

    # Save
    out_path = os.path.join(OUT, cat_dir, filename + ".xlsx")
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    return True

print("\n" + "=" * 60)
print("Generating all 240 complex templates...")
print("This will take a minute — each template has rich formatting, formulas & validation.")
print("=" * 60)

success = 0
errors = []
for i, item in enumerate(ALL_TEMPLATES):
    try:
        generate_template(*item)
        success += 1
        if success % 30 == 0:
            print(f"  [OK] {success}/240 templates generated...")
    except Exception as e:
        errors.append((item[1] if len(item) > 1 else "?", str(e)))
        print(f"  [ERR] Template {item[1] if len(item) > 1 else '?'}: {e}")

print("=" * 60)
print(f"\n[OK] Generation complete: {success}/{len(ALL_TEMPLATES)} templates created!")
if errors:
    print(f"[WARN] {len(errors)} errors:")
    for name, err in errors[:10]:
        print(f"  - {name}: {err}")
print(f"[OK] Output directory: {OUT}")
print(f"[OK] Every template includes: formulas, data validation, conditional formatting")
print(f"[OK] 3 sheets per template (Data + Instructions + About)")
print(f"\nNext: Open any .xlsx file in Excel, Google Sheets, or LibreOffice to explore!")
